import json
import inspect
from datetime import datetime
from pathlib import Path

import frappe
from frappe.tests.utils import FrappeTestCase

from kunal_enterprises.api import order_controls as owner_admin_order_controls
from kunal_enterprises.api.customer_access import status
from kunal_enterprises.api.branch_orders import mark_processing as branch_mark_processing
from kunal_enterprises.api.branch_orders import visible_orders as branch_visible_orders
from kunal_enterprises.api.health import smoke
from kunal_enterprises.api.utils import create_success_response, handle_error_response
from kunal_enterprises.api.sync_admin import (
	import_stock_excel_now,
	run_reconciliation_now,
	sync_masters_now,
	sync_stock_now,
	sync_vouchers_now,
)
from kunal_enterprises.api.order_controls import cancel_order, partially_close_order, resolve_manual_review
from kunal_enterprises.api.otp import resend_otp, send_otp, start_customer_signup, verify_customer_otp, verify_sales_employee_otp
from kunal_enterprises.api.product_groups import allowed as allowed_product_groups
from kunal_enterprises.api.product_groups import item_stock
from kunal_enterprises.api.product_groups import items as allowed_items
from kunal_enterprises.api.sales_employees import allowed_customers
from kunal_enterprises.api.token_verification import current_session, issue_token, revoke_token
from kunal_enterprises.kunal_enterprises.doctype.customer.customer import (
	search_tally_customer_ledgers,
	set_customer_client_code,
)
from kunal_enterprises.kunal_enterprises.doctype.tally_customer_ledger.tally_customer_ledger import get_mapped_customer
from kunal_enterprises.cron.tally_sync import sync_stock_snapshots
from kunal_enterprises.cron.tally_sync import sync_tally_masters
from kunal_enterprises.cron.tally_sync import sync_tally_vouchers
from kunal_enterprises.integrations.tally_postgres import _build_dev_stock_snapshot_rows, seed_dev_stock_snapshots
from kunal_enterprises import hooks
from kunal_enterprises.permission_query_conditions.orders import get_permission_query_conditions as order_permission_query
from kunal_enterprises.permission_query_conditions.orders import has_permission as order_has_permission
from kunal_enterprises.api.orders import detail as order_detail
from kunal_enterprises.api.orders import _financial_year_period
from kunal_enterprises.api.orders import history as order_history
from kunal_enterprises.api.orders import submit as submit_order
from kunal_enterprises.api.profile import get_profile, update_customer_profile
from kunal_enterprises.cron.reconciliation import run_reconciliation


class TestFoundation(FrappeTestCase):
	def test_smoke_reports_postgres_and_required_apps(self):
		response = smoke()

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["app"], "kunal_enterprises")
		self.assertEqual(response["data"]["database_type"], frappe.conf.get("db_type"))
		self.assertIn("frappe", response["data"]["installed_apps"])
		self.assertTrue(response["data"]["checks"]["custom_app_installed"])
		self.assertEqual(
			response["data"]["checks"]["frappe_whatsapp_installed"],
			"frappe_whatsapp" in response["data"]["installed_apps"],
		)
		self.assertEqual(frappe.db.sql("select 1")[0][0], 1)

	def test_response_envelopes_set_frappe_http_status_code(self):
		success = create_success_response("Created", {"name": "KE-TEST"}, status_code=201)
		self.assertTrue(success["success"])
		self.assertEqual(success["status"], "success")
		self.assertEqual(success["http_status_code"], 201)
		self.assertEqual(frappe.local.response["http_status_code"], 201)

		error = handle_error_response(
			frappe.PermissionError("Not allowed"),
			"Permission denied",
			log_title="Test Permission Response",
		)
		self.assertFalse(error["success"])
		self.assertEqual(error["status"], "error")
		self.assertEqual(error["http_status_code"], 403)
		self.assertEqual(frappe.local.response["http_status_code"], 403)
		self.assertEqual(error["error"]["message"], "Not allowed")

	def test_required_portal_roles_are_installed(self):
		for role in ("Owner", "Admin", "Branch Manager", "Branch Employee"):
			self.assertTrue(frappe.db.exists("Role", role), role)

	def test_required_goal_doctypes_are_installed(self):
		required_doctypes = (
			"Customer",
			"Mobile OTP",
			"Mobile Auth Token",
			"Sales Employee",
			"Sales Employee Assigned Customer",
			"Customer Product Group Access",
			"Sales Employee Product Group Access",
			"Tally Stock Group",
			"Tally Item",
			"Tally Stock Category",
			"Tally Godown",
			"Portal Branch",
			"Branch Godown Mapping",
			"Tally Unit",
			"Tally Customer Ledger",
			"Tally Stock Snapshot",
			"Tally Voucher",
			"Tally Voucher Line",
			"Tally Sync Run",
			"Tally Sync Error",
			"Order",
			"Order Item",
			"Order Godown Allocation",
			"Order Status Log",
			"Order PDF",
			"Order WhatsApp Notification",
			"Order Reconciliation Log",
			"Order Reference Sequence",
		)

		for doctype in required_doctypes:
			self.assertTrue(frappe.db.exists("DocType", doctype), doctype)

	def test_godown_fields_link_to_tally_godown_master(self):
		for doctype in (
			"Tally Stock Snapshot",
			"Tally Voucher Line",
			"Order Godown Allocation",
			"Branch Godown Mapping",
		):
			field = frappe.get_meta(doctype).get_field("godown")
			self.assertEqual(field.fieldtype, "Link", doctype)
			self.assertEqual(field.options, "Tally Godown", doctype)
			self.assertTrue(field.reqd, doctype)

	def test_manual_review_reconciliation_log_requires_reason_code_and_message(self):
		for missing_field in ("reason_code", "message"):
			payload = {
				"doctype": "Order Reconciliation Log",
				"status": "Manual Review",
				"reason_code": "CUSTOMER_CLIENT_CODE_MISMATCH",
				"message": "Customer Client Code mismatch",
				"created_at": "2026-05-19 12:00:00",
			}
			payload[missing_field] = ""
			with self.assertRaises(frappe.ValidationError):
				frappe.get_doc(payload).insert(ignore_permissions=True)

	def test_owner_admin_desk_permissions_exist_for_core_portal_doctypes(self):
		required_permissions = {
			"Customer": ("Owner", "Admin"),
			"Sales Employee": ("Owner", "Admin"),
			"Tally Customer Ledger": ("Owner", "Admin"),
			"Tally Stock Group": ("Owner", "Admin"),
			"Tally Item": ("Owner", "Admin"),
			"Tally Stock Category": ("Owner", "Admin"),
			"Tally Godown": ("Owner", "Admin"),
			"Tally Unit": ("Owner", "Admin"),
			"Tally Stock Snapshot": ("Owner", "Admin"),
			"Tally Voucher": ("Owner", "Admin"),
			"Tally Sync Run": ("Owner", "Admin"),
			"Tally Sync Error": ("Owner", "Admin"),
			"Order": ("Owner", "Admin", "Branch Manager", "Branch Employee"),
			"Order PDF": ("Owner", "Admin"),
			"Order WhatsApp Notification": ("Owner", "Admin"),
			"Order Reconciliation Log": ("Owner", "Admin"),
		}

		for doctype, roles in required_permissions.items():
			meta = frappe.get_meta(doctype)
			permission_roles = {permission.role for permission in meta.permissions}
			for role in roles:
				self.assertIn(role, permission_roles, f"{doctype} missing {role} permission")

	def test_required_role_profiles_are_installed_without_extra_roles(self):
		expected_profiles = {
			"Owner": {"Owner"},
			"Admin": {"Admin"},
			"Branch Manager": {"Branch Manager"},
			"Branch Employee": {"Branch Employee"},
		}

		for profile, expected_roles in expected_profiles.items():
			self.assertTrue(frappe.db.exists("Role Profile", profile), profile)
			role_profile = frappe.get_doc("Role Profile", profile)
			self.assertEqual(role_profile.role_profile, profile)
			self.assertEqual({row.role for row in role_profile.roles}, expected_roles)
			self.assertNotIn("System Manager", {row.role for row in role_profile.roles})

	def test_hooks_export_only_kunal_roles_and_role_profiles(self):
		fixture_filters = {fixture["dt"]: fixture.get("filters") for fixture in hooks.fixtures}
		expected_roles = ["Owner", "Admin", "Branch Manager", "Branch Employee"]

		self.assertIn("Role", fixture_filters)
		self.assertIn("Role Profile", fixture_filters)
		self.assertEqual(fixture_filters["Role"], [["role_name", "in", expected_roles]])
		self.assertEqual(fixture_filters["Role Profile"], [["name", "in", expected_roles]])

	def test_branch_roles_only_have_direct_order_read_until_row_hooks_exist(self):
		branch_roles = {"Branch Manager", "Branch Employee"}
		unsafe_without_row_hooks = ("Portal Branch", "Branch Godown Mapping", "Order Status Log")

		for doctype in unsafe_without_row_hooks:
			permissions = self._permissions_for(doctype)
			read_roles = {role for role, permission in permissions.items() if permission.read}
			self.assertTrue(
				branch_roles.isdisjoint(read_roles),
				f"{doctype} should not grant direct branch-role read without row-level hooks",
			)

		order_permissions = self._permissions_for("Order")
		for role in branch_roles:
			self.assertTrue(order_permissions[role].read)
			self.assertFalse(order_permissions[role].write)
			self.assertFalse(order_permissions[role].create)
			self.assertFalse(order_permissions[role].delete)

	def test_owner_admin_can_manage_branch_setup_doctypes(self):
		for doctype in ("Portal Branch", "Branch Godown Mapping"):
			permissions = self._permissions_for(doctype)
			for role in ("Owner", "Admin"):
				self.assertTrue(permissions[role].read, f"{doctype} missing {role} read")
				self.assertTrue(permissions[role].write, f"{doctype} missing {role} write")
				self.assertTrue(permissions[role].create, f"{doctype} missing {role} create")
				self.assertTrue(permissions[role].delete, f"{doctype} missing {role} delete")

	def test_tally_derived_doctypes_are_read_only_for_owner_admin(self):
		tally_derived_doctypes = (
			"Tally Customer Ledger",
			"Tally Stock Group",
			"Tally Item",
			"Tally Stock Category",
			"Tally Godown",
			"Tally Unit",
			"Tally Stock Snapshot",
			"Tally Voucher",
			"Tally Sync Run",
			"Tally Sync Error",
		)

		for doctype in tally_derived_doctypes:
			permissions = self._permissions_for(doctype)
			for role in ("Owner", "Admin"):
				self.assertTrue(permissions[role].read, f"{doctype} missing {role} read")
				self.assertFalse(permissions[role].write, f"{doctype} should not grant {role} write")
				self.assertFalse(permissions[role].create, f"{doctype} should not grant {role} create")
				self.assertFalse(permissions[role].delete, f"{doctype} should not grant {role} delete")

	def test_sensitive_internal_doctypes_are_not_exposed_to_branch_or_admin_roles(self):
		no_desk_doctypes = (
			"Mobile OTP",
			"Mobile Auth Token",
			"Order Reference Sequence",
		)
		operational_read_only_doctypes = (
			"Tally Sync Run",
			"Tally Sync Error",
			"Tally Voucher",
			"Order WhatsApp Notification",
			"Order Reconciliation Log",
		)
		for doctype in no_desk_doctypes:
			permissions = self._permissions_for(doctype)
			self.assertEqual(set(permissions), {"System Manager"}, doctype)
			for role in ("Owner", "Admin", "Branch Manager", "Branch Employee", "Guest", "All"):
				self.assertNotIn(role, permissions, f"{doctype} should not expose {role}")

		for doctype in operational_read_only_doctypes:
			permissions = self._permissions_for(doctype)
			for role in ("Owner", "Admin"):
				self.assertTrue(permissions[role].read, f"{doctype} missing {role} read")
				self.assertFalse(permissions[role].write, f"{doctype} should not grant {role} write")
				self.assertFalse(permissions[role].create, f"{doctype} should not grant {role} create")
				self.assertFalse(permissions[role].delete, f"{doctype} should not grant {role} delete")
			for role in ("Branch Manager", "Branch Employee", "Guest", "All"):
				self.assertNotIn(role, permissions, f"{doctype} should not expose {role}")

		self.assertEqual(frappe.get_meta("Tally Voucher Line").istable, 1)
		self.assertEqual(self._permissions_for("Tally Voucher Line"), {})

	def test_sync_admin_apis_do_not_allow_guest_access(self):
		admin_methods = (
			"kunal_enterprises.api.sync_admin.sync_masters_now",
			"kunal_enterprises.api.sync_admin.sync_stock_now",
			"kunal_enterprises.api.sync_admin.sync_vouchers_now",
			"kunal_enterprises.api.sync_admin.import_stock_excel_now",
			"kunal_enterprises.api.sync_admin.run_reconciliation_now",
		)
		for method in admin_methods:
			method_fn = frappe.get_attr(method)
			self.assertTrue(method_fn in frappe.whitelisted, method)
			self.assertFalse(method_fn in frappe.guest_methods, method)

	def _permissions_for(self, doctype):
		return {permission.role: permission for permission in frappe.get_meta(doctype).permissions}

	def test_mobile_token_api_endpoints_allow_guest_for_custom_auth_token_verification(self):
		mobile_token_methods = (
			"kunal_enterprises.api.customer_access.status",
			"kunal_enterprises.api.product_groups.allowed",
			"kunal_enterprises.api.product_groups.items",
			"kunal_enterprises.api.product_groups.item_stock",
			"kunal_enterprises.api.sales_employees.allowed_customers",
			"kunal_enterprises.api.orders.submit",
			"kunal_enterprises.api.orders.history",
			"kunal_enterprises.api.orders.detail",
			"kunal_enterprises.api.profile.get_profile",
			"kunal_enterprises.api.profile.update_customer_profile",
			"kunal_enterprises.api.token_verification.current_session",
			"kunal_enterprises.api.token_verification.revoke_token",
		)

		for method in mobile_token_methods:
			method_fn = frappe.get_attr(method)
			self.assertTrue(
				method_fn in frappe.guest_methods,
				f"{method} must allow guest so custom Auth-Token verification can run",
			)

	def test_order_facing_doctypes_do_not_expose_monetary_fields(self):
		order_facing_doctypes = (
			"Order",
			"Order Item",
			"Order Godown Allocation",
			"Order PDF",
			"Order WhatsApp Notification",
		)
		forbidden_terms = ("price", "rate", "amount", "tax", "discount", "value", "currency")
		violations = []

		for doctype in order_facing_doctypes:
			for field in frappe.get_meta(doctype).fields:
				text_parts = " ".join(
					filter(
						None,
						(
							field.fieldname,
							field.label,
							field.options if isinstance(field.options, str) else None,
						),
					)
				).lower()
				tokens = set(text_parts.replace("_", " ").replace("-", " ").split())
				if any(term in tokens for term in forbidden_terms):
					violations.append(f"{doctype}.{field.fieldname}")

		self.assertEqual(violations, [])

	def test_order_permission_query_hooks_are_registered(self):
		self.assertEqual(
			hooks.permission_query_conditions["Order"],
			"kunal_enterprises.permission_query_conditions.orders.get_permission_query_conditions",
		)
		self.assertEqual(
			hooks.has_permission["Order"],
			"kunal_enterprises.permission_query_conditions.orders.has_permission",
		)

class TestCustomerAppAccess(FrappeTestCase):
	def tearDown(self):
		frappe.db.rollback()

	def test_customer_form_uses_disable_label_for_active_customer_action(self):
		script = (
			Path(__file__).parents[1]
			/ "kunal_enterprises"
			/ "doctype"
			/ "customer"
			/ "customer.js"
		).read_text()

		self.assertIn('frm.doc.status === "Active"', script)
		self.assertIn('__("Disable")', script)
		self.assertIn('__("Disable Customer?")', script)
		self.assertIn('"disable_customer"', script)

	def test_disable_customer_removes_app_access_without_removing_admin_approval(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "DISABLE-CUSTOMER-001",
				"ledger_name": "Disable Customer Business",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Disable Customer",
				"business_legal_name": "Disable Customer Business",
				"mobile_number": "9000000041",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "DISABLE-CUSTOMER-001",
			}
		).insert()

		response = frappe.get_attr(
			"kunal_enterprises.kunal_enterprises.doctype.customer.customer.disable_customer"
		)(customer.name)
		customer.reload()

		self.assertEqual(customer.status, "Disabled")
		self.assertTrue(customer.admin_approved)
		self.assertFalse(customer.customer_app_access)
		self.assertFalse(response["customer_app_access"])
		self.assertIn("account_active", response["missing_requirements"])

	def test_customer_signup_stores_pending_details_in_otp_and_blocks_rejected_mobile_reuse(self):
		response = start_customer_signup(
			{
				"customer_name": "Signup Customer",
				"business_legal_name": "Signup Business",
				"gstin": "27ABCDE1234F1Z5",
				"mobile_number": " 9000000005 ",
				"email_id": "signup@example.com",
				"date_of_birth": "1990-01-02",
				"date_of_anniversary": "2015-03-04",
			}
		)
		self.assertTrue(response["success"])
		self.assertFalse(frappe.db.exists("Customer", {"mobile_number": "9000000005"}))
		self.assertEqual(response["data"]["status"], "Pending OTP")
		otp = frappe.get_doc(
			"Mobile OTP",
			{
				"mobile_number": "9000000005",
				"purpose": "Customer Signup",
				"status": "Open",
			},
		)
		self.assertEqual(otp.provider, "frappe_whatsapp")
		self.assertEqual(otp.provider_status, "Queued")
		self.assertEqual(otp.otp_type, "Account Creation")
		self.assertRegex(otp.otp_code, r"^\d{4}$")
		pending_payload = frappe.parse_json(otp.pending_payload)
		self.assertEqual(pending_payload["customer_name"], "Signup Customer")
		self.assertEqual(pending_payload["mobile_number"], "9000000005")
		self.assertEqual(pending_payload["email_id"], "signup@example.com")
		self.assertIn("Customer Signup", otp.request_payload)
		self.assertIn("Account Creation", otp.request_payload)
		self.assertIn("Queued", otp.provider_response)

		frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Rejected Signup Customer",
				"business_legal_name": "Rejected Signup Business",
				"mobile_number": "9000000030",
				"status": "Rejected",
			}
		).insert(ignore_permissions=True)

		retry = start_customer_signup(
			{
				"customer_name": "Retry Customer",
				"business_legal_name": "Retry Business",
				"mobile_number": "9000000030",
			}
		)

		self.assertFalse(retry["success"])
		self.assertIn("not available", retry["error"]["message"])

	def test_resend_customer_otp_is_blocked_until_cooldown_expires(self):
		customer_response = start_customer_signup(
			{
				"customer_name": "Resend Customer",
				"business_legal_name": "Resend Business",
				"mobile_number": "9000000021",
			}
		)
		resend_response = resend_otp("9000000021", "Customer")

		self.assertTrue(customer_response["success"])
		self.assertFalse(resend_response["success"])
		self.assertIn("wait", resend_response["error"]["message"].lower())

	def test_otp_verification_creates_customer_in_pending_admin_review_without_access(self):
		start_customer_signup(
			{
				"customer_name": "OTP Customer",
				"business_legal_name": "OTP Business",
				"mobile_number": "9000000006",
			}
		)
		frappe.get_doc(
			{
				"doctype": "Mobile OTP",
				"mobile_number": "9000000006",
				"otp_code": "1234",
				"purpose": "Customer Signup",
				"otp_type": "Account Creation",
				"pending_payload": json.dumps(
					{
						"customer_name": "OTP Customer",
						"business_legal_name": "OTP Business",
						"mobile_number": "9000000006",
					},
					sort_keys=True,
				),
				"status": "Open",
			}
		).insert(ignore_permissions=True)

		response = verify_customer_otp("9000000006", "1234")
		customer = frappe.get_doc("Customer", response["data"]["customer"])

		self.assertTrue(response["success"])
		self.assertRegex(customer.name, r"^KE-CUST-\d{5}$")
		self.assertEqual(customer.mobile_number, "9000000006")
		self.assertEqual(customer.status, "Pending Admin Review")
		self.assertTrue(customer.mobile_verified)
		self.assertTrue(customer.mobile_verified_at)
		self.assertFalse(customer.customer_app_access)

	def test_customer_signup_replaces_pending_payload_and_failed_creation_does_not_verify_otp(self):
		frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Existing GSTIN Customer",
				"business_legal_name": "Existing GSTIN Business",
				"mobile_number": "9000000036",
				"gstin": "27ABCDE1234F1Z6",
			}
		).insert()
		first_signup = start_customer_signup(
			{
				"customer_name": "Old Pending Customer",
				"business_legal_name": "Old Pending Business",
				"mobile_number": "9000000037",
				"email_id": "old-pending@example.com",
			}
		)
		rejected_signup = start_customer_signup(
			{
				"customer_name": "Rejected Pending Customer",
				"business_legal_name": "Rejected Pending Business",
				"mobile_number": "9000000037",
				"gstin": "27ABCDE1234F1Z6",
			}
		)
		replacement_signup = start_customer_signup(
			{
				"customer_name": "Replacement Pending Customer",
				"business_legal_name": "Replacement Pending Business",
				"mobile_number": "9000000037",
				"email_id": "replacement-pending@example.com",
			}
		)

		self.assertTrue(first_signup["success"])
		self.assertFalse(rejected_signup["success"])
		self.assertIn("GSTIN", rejected_signup["error"]["message"])
		self.assertTrue(replacement_signup["success"])

		open_otp = frappe.get_doc("Mobile OTP", replacement_signup["data"]["pending_otp"])
		pending_payload = frappe.parse_json(open_otp.pending_payload)
		self.assertEqual(pending_payload["customer_name"], "Replacement Pending Customer")
		self.assertEqual(pending_payload["email_id"], "replacement-pending@example.com")

		response = verify_customer_otp("9000000037", open_otp.otp_code)
		customer = frappe.get_doc("Customer", response["data"]["customer"])
		self.assertTrue(response["success"])
		self.assertEqual(customer.customer_name, "Replacement Pending Customer")
		self.assertEqual(customer.email_id, "replacement-pending@example.com")

		failing_otp = frappe.get_doc(
			{
				"doctype": "Mobile OTP",
				"mobile_number": "9000000038",
				"otp_code": "6789",
				"purpose": "Customer Signup",
				"otp_type": "Account Creation",
				"pending_payload": json.dumps(
					{
						"customer_name": "Duplicate GSTIN Pending Customer",
						"business_legal_name": "Duplicate GSTIN Pending Business",
						"mobile_number": "9000000038",
						"gstin": "27ABCDE1234F1Z6",
					},
					sort_keys=True,
				),
				"status": "Open",
			}
		).insert(ignore_permissions=True)
		failed_verify = verify_customer_otp("9000000038", "6789")
		failing_otp.reload()

		self.assertFalse(failed_verify["success"])
		self.assertEqual(failing_otp.status, "Open")
		self.assertFalse(failing_otp.verified_at)

	def test_customer_token_is_rejected_after_client_code_access_is_removed(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "AUTH-CUSTOMER-001",
				"ledger_name": "Auth Customer",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Auth Customer",
				"business_legal_name": "Auth Business",
				"mobile_number": "9000000007",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "AUTH-CUSTOMER-001",
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Mobile OTP",
				"mobile_number": "9000000007",
				"otp_code": "6543",
				"purpose": "Customer Signup",
				"status": "Open",
			}
		).insert(ignore_permissions=True)

		verified = verify_customer_otp("9000000007", "6543")
		session = current_session({"Auth-Token": f"Bearer {verified['data']['access_token']}"})
		frappe.db.set_value("Customer", customer.name, "client_code", None)
		rejected = current_session({"Auth-Token": f"Bearer {verified['data']['access_token']}"})

		self.assertTrue(verified["success"])
		self.assertEqual(verified["data"]["identity_type"], "Customer")
		self.assertEqual(session["data"]["customer"], customer.name)
		self.assertFalse(rejected["success"])
		self.assertIn("Invalid or inactive token", rejected["error"]["message"])

	def test_logout_revokes_customer_token(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "LOGOUT-CUSTOMER-001",
				"ledger_name": "Logout Customer",
				"is_active": 1,
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Logout Customer",
				"business_legal_name": "Logout Business",
				"mobile_number": "9000000008",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "LOGOUT-CUSTOMER-001",
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Mobile OTP",
				"mobile_number": "9000000008",
				"otp_code": "7777",
				"purpose": "Customer Signup",
				"status": "Open",
			}
		).insert(ignore_permissions=True)
		verified = verify_customer_otp("9000000008", "7777")
		headers = {"Auth-Token": f"Bearer {verified['data']['access_token']}"}

		logout = revoke_token(headers)
		rejected = current_session(headers)

		self.assertTrue(logout["success"])
		self.assertEqual(frappe.db.get_value("Mobile Auth Token", verified["data"]["token"], "status"), "Revoked")
		self.assertFalse(rejected["success"])
		self.assertIn("Invalid or inactive token", rejected["error"]["message"])

	def test_session_and_logout_read_auth_token_from_request_headers(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "REQUEST-HEADER-CUSTOMER-001",
				"ledger_name": "Request Header Customer",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Request Header Customer",
				"business_legal_name": "Request Header Business",
				"mobile_number": "9000000032",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "REQUEST-HEADER-CUSTOMER-001",
			}
		).insert()
		token = issue_token("Customer", customer.name)
		previous_request = getattr(frappe.local, "request", None)
		frappe.local.request = frappe._dict(
			{"headers": {"Auth-Token": f"Bearer {token['access_token']}"}}
		)
		try:
			session = current_session()
			logout = revoke_token()
			rejected = current_session()
		finally:
			frappe.local.request = previous_request

		self.assertTrue(session["success"])
		self.assertEqual(session["data"]["customer"], customer.name)
		self.assertTrue(logout["success"])
		self.assertEqual(logout["data"]["status"], "Revoked")
		self.assertFalse(rejected["success"])

	def test_customer_app_access_requires_valid_tally_client_code(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "KE-CLIENT-001",
				"ledger_name": "Kunal Test Business",
				"is_active": 1,
			}
		).insert()

		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Kunal Test Customer",
				"business_legal_name": "Kunal Test Business",
				"mobile_number": "9000000001",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "KE-CLIENT-001",
			}
		).insert()

		response = status(customer.name)

		self.assertTrue(response["success"])
		self.assertTrue(response["data"]["customer_app_access"])
		self.assertEqual(response["data"]["missing_requirements"], [])

	def test_customer_without_client_code_has_no_app_access(self):
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "No Code Customer",
				"business_legal_name": "No Code Business",
				"mobile_number": "9000000002",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
			}
		).insert()

		response = status(customer.name)

		self.assertTrue(response["success"])
		self.assertFalse(response["data"]["customer_app_access"])
		self.assertIn("client_code_present", response["data"]["missing_requirements"])

	def test_customer_access_status_requires_matching_customer_token_when_headers_are_supplied(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "ACCESS-TOKEN-001",
				"ledger_name": "Access Token Business",
				"is_active": 1,
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "ACCESS-TOKEN-002",
				"ledger_name": "Other Access Token Business",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Access Token Customer",
				"business_legal_name": "Access Token Business",
				"mobile_number": "9000000030",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "ACCESS-TOKEN-001",
			}
		).insert()
		other_customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Other Access Token Customer",
				"business_legal_name": "Other Access Token Business",
				"mobile_number": "9000000031",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "ACCESS-TOKEN-002",
			}
		).insert()
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_token_response = status(customer.name, headers={})
		mismatched_token_response = status(
			customer.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_response = status(
			customer.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_token_response["success"])
		self.assertIn("token identity", mismatched_token_response["error"]["message"])
		self.assertTrue(valid_response["success"])
		self.assertTrue(valid_response["data"]["customer_app_access"])

	def test_customer_rejects_client_code_missing_from_tally_ledgers(self):
		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Customer",
					"customer_name": "Invalid Code Customer",
					"business_legal_name": "Invalid Code Business",
					"mobile_number": "9000000003",
					"mobile_verified": 1,
					"admin_approved": 1,
					"status": "Active",
					"client_code": "MISSING-CODE",
				}
			).insert()

	def test_customer_client_code_is_unique_and_removal_disables_app_access(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "UNIQUE-CUSTOMER-001",
				"ledger_name": "Unique Customer Business",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Unique Code Customer",
				"business_legal_name": "Unique Customer Business",
				"mobile_number": "9000000024",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "UNIQUE-CUSTOMER-001",
			}
		).insert()

		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Customer",
					"customer_name": "Duplicate Code Customer",
					"business_legal_name": "Duplicate Code Business",
					"mobile_number": "9000000025",
					"mobile_verified": 1,
					"admin_approved": 1,
					"status": "Active",
					"client_code": "UNIQUE-CUSTOMER-001",
				}
			).insert()

		customer.client_code = None
		customer.save(ignore_permissions=True)
		response = status(customer.name)

		customer.reload()
		self.assertFalse(customer.customer_app_access)
		self.assertFalse(response["data"]["customer_app_access"])
		self.assertIn("client_code_present", response["data"]["missing_requirements"])

	def test_customer_client_code_can_be_set_through_controlled_action(self):
		for client_code in ("BUTTON-CODE-001", "BUTTON-CODE-002"):
			frappe.get_doc(
				{
					"doctype": "Tally Customer Ledger",
					"client_code": client_code,
					"ledger_name": f"Ledger {client_code}",
					"is_active": 1,
				}
			).insert()

		first_customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Client Code Button Customer",
				"business_legal_name": "Client Code Button Business",
				"mobile_number": "9000000039",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
			}
		).insert()
		second_customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Duplicate Button Customer",
				"business_legal_name": "Duplicate Button Business",
				"mobile_number": "9000000040",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "BUTTON-CODE-002",
			}
		).insert()

		response = set_customer_client_code(first_customer.name, " BUTTON-CODE-001 ")
		first_customer.reload()

		self.assertEqual(first_customer.client_code, "BUTTON-CODE-001")
		self.assertTrue(response["customer_app_access"])

		with self.assertRaises(frappe.ValidationError):
			set_customer_client_code(first_customer.name, second_customer.client_code)

	def test_tally_ledger_search_only_returns_unassigned_ledgers(self):
		for client_code, ledger_name in (
			("SEARCH-CODE-001", "Search Available Ledger"),
			("SEARCH-CODE-002", "Search Assigned Ledger"),
			("SEARCH-CODE-003", "Search Inactive Ledger"),
		):
			frappe.get_doc(
				{
					"doctype": "Tally Customer Ledger",
					"client_code": client_code,
					"ledger_name": ledger_name,
					"is_active": 0 if client_code == "SEARCH-CODE-003" else 1,
				}
			).insert()

		assigned_customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Search Assigned Customer",
				"business_legal_name": "Search Assigned Business",
				"mobile_number": "9000000041",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "SEARCH-CODE-002",
			}
		).insert()
		current_customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Search Current Customer",
				"business_legal_name": "Search Current Business",
				"mobile_number": "9000000042",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
			}
		).insert()

		results = search_tally_customer_ledgers("Search", current_customer.name)
		client_codes = {row.client_code for row in results}

		self.assertIn("SEARCH-CODE-001", client_codes)
		self.assertNotIn("SEARCH-CODE-002", client_codes)
		self.assertNotIn("SEARCH-CODE-003", client_codes)

		own_results = search_tally_customer_ledgers("SEARCH-CODE-002", assigned_customer.name)
		own_result_by_code = {row.client_code: row for row in own_results}
		self.assertIn("SEARCH-CODE-002", own_result_by_code)
		self.assertEqual(own_result_by_code["SEARCH-CODE-002"].mapped_customer, assigned_customer.name)

	def test_tally_ledger_reports_mapped_customer(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "MAPPED-CODE-001",
				"ledger_name": "Mapped Ledger",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Mapped Customer",
				"business_legal_name": "Mapped Business",
				"mobile_number": "9000000043",
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": "MAPPED-CODE-001",
			}
		).insert()

		mapping = get_mapped_customer("MAPPED-CODE-001")

		self.assertEqual(mapping.name, customer.name)
		self.assertEqual(mapping.business_legal_name, "Mapped Business")

	def test_customer_email_and_gstin_are_unique_when_present(self):
		frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Unique Identity Customer",
				"business_legal_name": "Unique Identity Business",
				"mobile_number": "9000000031",
				"email_id": "  UNIQUE@example.com ",
				"gstin": " 27abcde1234f1z5 ",
			}
		).insert()

		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Customer",
					"customer_name": "Duplicate Email Customer",
					"business_legal_name": "Duplicate Email Business",
					"mobile_number": "9000000032",
					"email_id": "unique@example.com",
				}
			).insert()

		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Customer",
					"customer_name": "Duplicate GSTIN Customer",
					"business_legal_name": "Duplicate GSTIN Business",
					"mobile_number": "9000000033",
					"gstin": "27ABCDE1234F1Z5",
				}
			).insert()

	def test_customer_and_sales_employee_use_locked_system_generated_ids(self):
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Series Customer",
				"business_legal_name": "Series Business",
				"mobile_number": "9000000034",
			}
		).insert()
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Series Sales Employee",
				"mobile_number": "9000000035",
				"status": "Active",
			}
		).insert()

		self.assertRegex(customer.name, r"^KE-CUST-\d{5}$")
		self.assertEqual(customer.mobile_number, "9000000034")
		self.assertRegex(sales_employee.name, r"^KE-SE-\d{4}$")
		self.assertEqual(sales_employee.mobile_number, "9000000035")
		self.assertFalse(frappe.get_meta("Customer").allow_rename)
		self.assertFalse(frappe.get_meta("Sales Employee").allow_rename)

	def test_mobile_number_is_unique_across_customers_and_sales_employees(self):
		frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Assigned Sales Employee",
				"mobile_number": "9000000004",
				"status": "Active",
			}
		).insert()

		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Customer",
					"customer_name": "Duplicate Mobile Customer",
					"business_legal_name": "Duplicate Mobile Business",
					"mobile_number": "9000000004",
				}
			).insert()

		frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Existing Mobile Customer",
				"business_legal_name": "Existing Mobile Business",
				"mobile_number": "9000000029",
			}
		).insert()

		with self.assertRaises(frappe.ValidationError):
			frappe.get_doc(
				{
					"doctype": "Sales Employee",
					"sales_employee_name": "Duplicate Mobile Sales Employee",
					"mobile_number": "9000000029",
					"status": "Active",
				}
			).insert()


class TestSalesEmployeeAuth(FrappeTestCase):
	def tearDown(self):
		frappe.db.rollback()

	def test_sales_employee_form_uses_disable_label_for_disabled_status_action(self):
		script = (
			Path(__file__).parents[1]
			/ "kunal_enterprises"
			/ "doctype"
			/ "sales_employee"
			/ "sales_employee.js"
		).read_text()

		self.assertIn('__("Disable")', script)
		self.assertIn('__("Disable Sales Employee?")', script)
		self.assertIn('"disable_sales_employee"', script)
		self.assertNotIn('__("Reject")', script)
		self.assertNotIn('__("Reject Sales Employee?")', script)
		self.assertNotIn('"reject_sales_employee"', script)

	def test_send_sales_employee_otp_creates_login_otp_for_active_employee(self):
		frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Send OTP Employee",
				"mobile_number": "9000000022",
				"status": "Active",
				"mobile_verified": 1,
			}
		).insert()

		response = send_otp("9000000022", "Sales Employee")

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["purpose"], "Sales Employee Login")
		self.assertEqual(response["data"]["otp_type"], "Login")
		self.assertFalse(response["data"]["mobile_verification_required"])
		self.assertEqual(response["data"]["cooldown_seconds"], 45)
		self.assertTrue(
			frappe.db.exists(
				"Mobile OTP",
				{
					"mobile_number": "9000000022",
					"purpose": "Sales Employee Login",
					"otp_type": "Login",
					"status": "Open",
				},
			)
		)

	def test_send_sales_employee_otp_requires_mobile_verification_for_first_login(self):
		frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "First Login Verification Employee",
				"mobile_number": "9000000024",
				"status": "Active",
				"mobile_verified": 0,
			}
		).insert()

		response = send_otp("9000000024", "Sales Employee")

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["purpose"], "Sales Employee Login")
		self.assertEqual(response["data"]["otp_type"], "Account Verification")
		self.assertTrue(response["data"]["mobile_verification_required"])
		self.assertTrue(
			frappe.db.exists(
				"Mobile OTP",
				{
					"mobile_number": "9000000024",
					"purpose": "Sales Employee Login",
					"otp_type": "Account Verification",
					"status": "Open",
				},
			)
		)

	def test_send_sales_employee_otp_blocks_disabled_employee(self):
		frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Disabled OTP Employee",
				"mobile_number": "9000000023",
				"status": "Disabled",
			}
		).insert()

		response = send_otp("9000000023", "Sales Employee")

		self.assertFalse(response["success"])
		self.assertIn("Disabled", response["error"]["message"])
		self.assertFalse(frappe.db.exists("Mobile OTP", {"mobile_number": "9000000023"}))

	def test_sales_employee_otp_login_is_blocked_when_disabled(self):
		active_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Active Login Employee",
				"mobile_number": "9000000011",
				"status": "Active",
			}
		).insert()
		disabled_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Disabled Login Employee",
				"mobile_number": "9000000012",
				"status": "Disabled",
			}
		).insert()
		self._create_otp("9000000011", "1111")
		self._create_otp("9000000012", "2222")

		active_response = verify_sales_employee_otp("9000000011", "1111")
		active_session = current_session({"Auth-Token": f"Bearer {active_response['data']['access_token']}"})
		disabled_response = verify_sales_employee_otp("9000000012", "2222")

		self.assertTrue(active_response["success"])
		self.assertEqual(active_response["data"]["sales_employee"], active_employee.name)
		self.assertEqual(active_response["data"]["identity_type"], "Sales Employee")
		self.assertTrue(active_response["data"]["mobile_verified"])
		self.assertTrue(active_response["data"]["verification_completed"])
		self.assertEqual(active_session["data"]["sales_employee"], active_employee.name)
		self.assertFalse(disabled_response["success"])
		self.assertIn("Disabled", disabled_response["error"]["message"])
		self.assertFalse(
			frappe.db.exists(
				"Mobile Auth Token",
				{
					"identity_type": "Sales Employee",
					"identity": disabled_employee.name,
					"status": "Active",
				},
			)
		)

	def _create_otp(self, mobile_number, otp_code):
		frappe.get_doc(
			{
				"doctype": "Mobile OTP",
				"mobile_number": mobile_number,
				"otp_code": otp_code,
				"purpose": "Sales Employee Login",
				"otp_type": "Login",
				"status": "Open",
			}
		).insert(ignore_permissions=True)


class TestSalesEmployeeCustomerAccess(FrappeTestCase):
	def tearDown(self):
		frappe.db.rollback()

	def test_allowed_customer_search_respects_assignments_and_hides_client_code(self):
		alpha = self._create_active_customer("9000000013", "SEARCH-ALPHA", "Alpha Customer", "Alpha Business")
		beta = self._create_active_customer("9000000014", "SEARCH-BETA", "Beta Customer", "Beta Business")
		disabled = self._create_active_customer("9000000015", "SEARCH-DISABLED", "Disabled Customer", "Disabled Business")
		frappe.db.set_value("Customer", disabled.name, "status", "Disabled")
		open_employee = self._create_sales_employee("Open Customer Employee", "9000000016")
		assigned_employee = self._create_sales_employee(
			"Assigned Customer Employee",
			"9000000017",
			assigned_customers=[beta.name],
		)

		open_response = allowed_customers(open_employee.name, search="SEARCH-ALPHA")
		assigned_response = allowed_customers(assigned_employee.name)

		self.assertTrue(open_response["success"])
		self.assertEqual([row["customer"] for row in open_response["data"]["customers"]], [alpha.name])
		self.assertNotIn("client_code", open_response["data"]["customers"][0])
		self.assertTrue(assigned_response["success"])
		self.assertEqual([row["customer"] for row in assigned_response["data"]["customers"]], [beta.name])

	def test_allowed_customers_requires_matching_sales_employee_token_when_headers_are_supplied(self):
		customer = self._create_active_customer("9000000026", "SEARCH-TOKEN", "Token Customer", "Token Business")
		sales_employee = self._create_sales_employee(
			"Token Customer Employee",
			"9000000027",
			assigned_customers=[customer.name],
		)
		other_employee = self._create_sales_employee("Other Token Customer Employee", "9000000028")
		valid_token = issue_token("Sales Employee", sales_employee.name)
		other_token = issue_token("Sales Employee", other_employee.name)

		missing_token_response = allowed_customers(sales_employee.name, headers={})
		mismatched_token_response = allowed_customers(
			sales_employee.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_response = allowed_customers(
			sales_employee.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_token_response["success"])
		self.assertIn("token identity", mismatched_token_response["error"]["message"])
		self.assertTrue(valid_response["success"])
		self.assertEqual([row["customer"] for row in valid_response["data"]["customers"]], [customer.name])

	def _create_active_customer(self, mobile_number, client_code, customer_name, business_name):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": client_code,
				"ledger_name": business_name,
				"is_active": 1,
			}
		).insert()
		return frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": customer_name,
				"business_legal_name": business_name,
				"mobile_number": mobile_number,
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": client_code,
			}
		).insert()

	def _create_sales_employee(self, employee_name, mobile_number, assigned_customers=None):
		return frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": employee_name,
				"mobile_number": mobile_number,
				"status": "Active",
				"assigned_customers": [
					{"customer": customer}
					for customer in (assigned_customers or [])
				],
			}
		).insert()


class TestMobileProfile(FrappeTestCase):
	def tearDown(self):
		frappe.db.rollback()

	def test_customer_profile_update_is_limited_and_sales_employee_profile_is_read_only(self):
		customer = self._create_active_customer()
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Read Only Profile Employee",
				"mobile_number": "9000000019",
				"email_id": "sales-original@example.com",
				"status": "Active",
			}
		).insert()

		customer_update = update_customer_profile(
			customer.name,
			{
				"email_id": "updated@example.com",
				"date_of_birth": "1991-02-03",
				"date_of_anniversary": "2016-04-05",
				"business_legal_name": "Mutated Business",
				"client_code": "PROFILE-HACK",
			},
		)
		customer_profile = get_profile("Customer", customer.name)
		sales_profile = get_profile("Sales Employee", sales_employee.name)

		customer.reload()
		self.assertTrue(customer_update["success"])
		self.assertEqual(customer.email_id, "updated@example.com")
		self.assertEqual(str(customer.date_of_birth), "1991-02-03")
		self.assertEqual(str(customer.date_of_anniversary), "2016-04-05")
		self.assertEqual(customer.business_legal_name, "Profile Business")
		self.assertEqual(customer.client_code, "PROFILE-CUSTOMER-001")
		self.assertNotIn("client_code", customer_profile["data"])
		self.assertEqual(
			customer_profile["data"]["editable_fields"],
			["email_id", "date_of_birth", "date_of_anniversary"],
		)
		self.assertFalse(sales_profile["data"]["editable_fields"])
		self.assertEqual(sales_profile["data"]["sales_employee"], sales_employee.name)

	def test_customer_profile_rejects_removed_app_access(self):
		customer = self._create_active_customer()
		customer.status = "Disabled"
		customer.save(ignore_permissions=True)

		profile_response = get_profile("Customer", customer.name)
		update_response = update_customer_profile(customer.name, {"email_id": "blocked@example.com"})

		customer.reload()
		self.assertFalse(profile_response["success"])
		self.assertFalse(update_response["success"])
		self.assertIn("Customer App Access", profile_response["error"]["message"])
		self.assertIn("Customer App Access", update_response["error"]["message"])
		self.assertEqual(customer.email_id, "original@example.com")

	def test_sales_employee_profile_rejects_disabled_employee(self):
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Disabled Profile Employee",
				"mobile_number": "9000000020",
				"email_id": "disabled-profile@example.com",
				"status": "Disabled",
			}
		).insert()

		response = get_profile("Sales Employee", sales_employee.name)

		self.assertFalse(response["success"])
		self.assertIn("Sales Employee", response["error"]["message"])

	def test_customer_profile_requires_matching_customer_token_when_headers_are_supplied(self):
		customer = self._create_active_customer()
		other_customer = self._create_active_customer(
			client_code="PROFILE-CUSTOMER-002",
			mobile_number="9000000029",
			customer_name="Other Profile Customer",
			business_name="Other Profile Business",
			email_id="other@example.com",
		)
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_token_response = get_profile("Customer", customer.name, headers={})
		mismatched_profile_response = get_profile(
			"Customer",
			customer.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		mismatched_update_response = update_customer_profile(
			customer.name,
			{"email_id": "blocked-token@example.com"},
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_profile_response = get_profile(
			"Customer",
			customer.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)
		valid_update_response = update_customer_profile(
			customer.name,
			{"email_id": "token-updated@example.com"},
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		customer.reload()
		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_profile_response["success"])
		self.assertIn("token identity", mismatched_profile_response["error"]["message"])
		self.assertFalse(mismatched_update_response["success"])
		self.assertIn("token identity", mismatched_update_response["error"]["message"])
		self.assertTrue(valid_profile_response["success"])
		self.assertTrue(valid_update_response["success"])
		self.assertEqual(customer.email_id, "token-updated@example.com")

	def _create_active_customer(
		self,
		client_code="PROFILE-CUSTOMER-001",
		mobile_number="9000000018",
		customer_name="Profile Customer",
		business_name="Profile Business",
		email_id="original@example.com",
	):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": client_code,
				"ledger_name": business_name,
				"is_active": 1,
			}
		).insert()
		return frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": customer_name,
				"business_legal_name": business_name,
				"mobile_number": mobile_number,
				"email_id": email_id,
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": client_code,
			}
		).insert()


class TestProductGroupAccess(FrappeTestCase):
	def tearDown(self):
		frappe.db.rollback()

	def test_customer_blank_product_group_filter_returns_all_active_root_groups(self):
		group_a = self._create_product_group("PG Blank A")
		group_b = self._create_product_group("PG Blank B")
		self._create_product_group("PG Blank Inactive", is_active=0)
		customer = self._create_active_customer("9000000101", "PG-BLANK-001")

		response = allowed_product_groups(customer.name)

		self.assertTrue(response["success"])
		group_names = {group["name"] for group in response["data"]["product_groups"]}
		self.assertIn(group_a.name, group_names)
		self.assertIn(group_b.name, group_names)
		self.assertNotIn("PG Blank Inactive", group_names)

	def test_customer_product_group_filter_limits_visible_groups_and_items(self):
		allowed_group = self._create_product_group("PG Customer Allowed")
		blocked_group = self._create_product_group("PG Customer Blocked")
		visible_item = self._create_item("Allowed Item", allowed_group.name)
		self._create_item("Blocked Item", blocked_group.name)
		customer = self._create_active_customer(
			"9000000102",
			"PG-CUSTOMER-001",
			product_groups=[allowed_group.name],
		)

		groups_response = allowed_product_groups(customer.name)
		items_response = allowed_items(customer.name, allowed_group.name)
		blocked_items_response = allowed_items(customer.name, blocked_group.name)

		self.assertTrue(groups_response["success"])
		self.assertEqual(
			[group["name"] for group in groups_response["data"]["product_groups"]],
			[allowed_group.name],
		)
		self.assertTrue(items_response["success"])
		self.assertEqual([item["name"] for item in items_response["data"]["items"]], [visible_item.name])
		self.assertFalse(blocked_items_response["success"])

	def test_allowed_items_show_total_from_godown_stock_snapshots(self):
		product_group = self._create_product_group("PG Snapshot Total")
		item = self._create_item("Snapshot Total Item", product_group.name)
		customer = self._create_active_customer(
			"9000000440",
			"PG-SNAPSHOT-TOTAL-001",
			product_groups=[product_group.name],
		)
		self._create_stock_snapshot(item.name, "Snapshot Godown A", 37)
		self._create_stock_snapshot(item.name, "Snapshot Godown B", 240)

		response = allowed_items(customer.name, product_group.name)

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["items"][0]["total_closing_balance"], 277)

	def test_sales_employee_product_group_access_intersects_with_customer_access(self):
		customer_group = self._create_product_group("PG Customer Only")
		shared_group = self._create_product_group("PG Shared")
		employee_group = self._create_product_group("PG Employee Only")
		customer = self._create_active_customer(
			"9000000103",
			"PG-INTERSECTION-001",
			product_groups=[customer_group.name, shared_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "PG Sales Employee",
				"mobile_number": "9000000104",
				"status": "Active",
				"product_group_access": [
					{"product_group": shared_group.name},
					{"product_group": employee_group.name},
				],
			}
		).insert()

		response = allowed_product_groups(customer.name, sales_employee.name)

		self.assertTrue(response["success"])
		self.assertEqual(
			[group["name"] for group in response["data"]["product_groups"]],
			[shared_group.name],
		)

	def test_sales_employee_customer_assignment_limits_customer_context(self):
		allowed_customer = self._create_active_customer("9000000105", "PG-ASSIGNED-001")
		blocked_customer = self._create_active_customer("9000000106", "PG-ASSIGNED-002")
		product_group = self._create_product_group("PG Assignment")
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Assigned PG Sales Employee",
				"mobile_number": "9000000107",
				"status": "Active",
				"assigned_customers": [{"customer": allowed_customer.name}],
				"product_group_access": [{"product_group": product_group.name}],
			}
		).insert()

		allowed_response = allowed_product_groups(allowed_customer.name, sales_employee.name)
		blocked_response = allowed_product_groups(blocked_customer.name, sales_employee.name)

		self.assertTrue(allowed_response["success"])
		self.assertFalse(blocked_response["success"])

	def test_item_stock_returns_advisory_godown_snapshots_for_allowed_item(self):
		product_group = self._create_product_group("Stock PG Allowed")
		item = self._create_item("Stock Item Allowed", product_group.name)
		customer = self._create_active_customer("9000000111", "PG-STOCK-001")
		self._create_stock_snapshot(item.name, "Main Godown", 12)
		self._create_stock_snapshot(item.name, "Zero Godown", 0)

		response = item_stock(customer.name, item.name)

		self.assertTrue(response["success"])
		self.assertTrue(response["data"]["stock_is_advisory"])
		self.assertEqual(
			{row["godown"]: row["quantity"] for row in response["data"]["godowns"]},
			{"Main Godown": 12, "Zero Godown": 0},
		)

	def test_product_item_and_stock_reads_require_matching_customer_token_when_headers_are_supplied(self):
		product_group = self._create_product_group("PG Token")
		item = self._create_item("Token Item", product_group.name)
		customer = self._create_active_customer("9000000430", "PG-TOKEN-001")
		other_customer = self._create_active_customer("9000000431", "PG-TOKEN-002")
		self._create_stock_snapshot(item.name, "Token Stock Godown", 6)
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_group_response = allowed_product_groups(customer.name, headers={})
		mismatched_items_response = allowed_items(
			customer.name,
			product_group.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		mismatched_stock_response = item_stock(
			customer.name,
			item.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_group_response = allowed_product_groups(
			customer.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)
		valid_items_response = allowed_items(
			customer.name,
			product_group.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)
		valid_stock_response = item_stock(
			customer.name,
			item.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_group_response["success"])
		self.assertEqual(missing_group_response["http_status_code"], 401)
		self.assertFalse(mismatched_items_response["success"])
		self.assertIn("token identity", mismatched_items_response["error"]["message"])
		self.assertFalse(mismatched_stock_response["success"])
		self.assertIn("token identity", mismatched_stock_response["error"]["message"])
		self.assertTrue(valid_group_response["success"])
		self.assertTrue(valid_items_response["success"])
		self.assertTrue(valid_stock_response["success"])

	def test_item_stock_rejects_item_outside_customer_product_group_access(self):
		allowed_group = self._create_product_group("Stock PG Visible")
		blocked_group = self._create_product_group("Stock PG Hidden")
		blocked_item = self._create_item("Stock Item Hidden", blocked_group.name)
		customer = self._create_active_customer(
			"9000000112",
			"PG-STOCK-002",
			product_groups=[allowed_group.name],
		)
		self._create_stock_snapshot(blocked_item.name, "Hidden Godown", 10)

		response = item_stock(customer.name, blocked_item.name)

		self.assertFalse(response["success"])

	def test_item_stock_rejects_customer_after_app_access_removed(self):
		product_group = self._create_product_group("Stock PG Disabled Customer")
		item = self._create_item("Stock Item Disabled Customer", product_group.name)
		customer = self._create_active_customer("9000000418", "PG-STOCK-DISABLED-001")
		self._create_stock_snapshot(item.name, "Disabled Customer Godown", 8)
		frappe.db.set_value("Customer", customer.name, "status", "Disabled")
		frappe.db.set_value("Customer", customer.name, "customer_app_access", 0)

		response = item_stock(customer.name, item.name)

		self.assertFalse(response["success"])
		self.assertIn("Customer App Access", response["error"]["message"])

	def test_item_stock_rejects_disabled_sales_employee_after_access_removed(self):
		product_group = self._create_product_group("Stock PG Disabled Sales")
		item = self._create_item("Stock Item Disabled Sales", product_group.name)
		customer = self._create_active_customer(
			"9000000419",
			"PG-STOCK-DISABLED-SALES-001",
			product_groups=[product_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Disabled Stock Sales Employee",
				"mobile_number": "9000000420",
				"status": "Disabled",
				"assigned_customers": [{"customer": customer.name}],
				"product_group_access": [{"product_group": product_group.name}],
			}
		).insert()
		self._create_stock_snapshot(item.name, "Disabled Sales Stock Godown", 8)

		response = item_stock(customer.name, item.name, sales_employee=sales_employee.name)

		self.assertFalse(response["success"])
		self.assertIn("Sales Employee", response["error"]["message"])

	def test_item_stock_rejects_sales_employee_unassigned_customer_context(self):
		product_group = self._create_product_group("Stock PG Unassigned Sales")
		item = self._create_item("Stock Item Unassigned Sales", product_group.name)
		allowed_customer = self._create_active_customer(
			"9000000421",
			"PG-STOCK-UNASSIGNED-001",
			product_groups=[product_group.name],
		)
		blocked_customer = self._create_active_customer(
			"9000000422",
			"PG-STOCK-UNASSIGNED-002",
			product_groups=[product_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Unassigned Stock Sales Employee",
				"mobile_number": "9000000423",
				"status": "Active",
				"assigned_customers": [{"customer": allowed_customer.name}],
				"product_group_access": [{"product_group": product_group.name}],
			}
		).insert()
		self._create_stock_snapshot(item.name, "Unassigned Sales Stock Godown", 8)

		response = item_stock(blocked_customer.name, item.name, sales_employee=sales_employee.name)

		self.assertFalse(response["success"])
		self.assertIn("assigned", response["error"]["message"])

	def _create_product_group(self, group_name, is_active=1):
		return frappe.get_doc(
			{
				"doctype": "Tally Stock Group",
				"group_name": group_name,
				"is_root": 1,
				"depth": 0,
				"full_path": group_name,
				"is_active": is_active,
			}
		).insert()

	def _create_item(self, item_name, root_stock_group):
		return frappe.get_doc(
			{
				"doctype": "Tally Item",
				"item_name": item_name,
				"root_stock_group": root_stock_group,
				"uom": "PCS",
				"total_closing_balance": 10,
				"is_active": 1,
			}
		).insert()

	def _create_stock_snapshot(self, item, godown, quantity):
		if not frappe.db.exists("Tally Godown", godown):
			frappe.get_doc(
				{
					"doctype": "Tally Godown",
					"godown_name": godown,
					"is_active": 1,
				}
			).insert()
		return frappe.get_doc(
			{
				"doctype": "Tally Stock Snapshot",
				"item": item,
				"godown": godown,
				"quantity": quantity,
				"uom": "PCS",
				"as_on_date": "2026-05-19",
				"source_company": "Kunal Test Company",
				"synced_at": "2026-05-19 12:00:00",
			}
		).insert()

	def _create_active_customer(self, mobile_number, client_code, product_groups=None):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": client_code,
				"ledger_name": f"Ledger {client_code}",
				"is_active": 1,
			}
		).insert()
		return frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": f"Customer {client_code}",
				"business_legal_name": f"Business {client_code}",
				"mobile_number": mobile_number,
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": client_code,
				"product_group_access": [
					{"product_group": product_group}
					for product_group in (product_groups or [])
				],
			}
		).insert()


class TestOrderSubmission(FrappeTestCase):
	def setUp(self):
		self._ensure_godowns(
			"Main Godown",
			"Immutable Godown",
			"Disabled Godown",
			"Token Godown",
			"Sales Godown",
			"Disabled Sales Godown",
			"Notify Godown",
			"History Token Godown",
			"Other Godown",
			"All History Godown 1",
			"All History Godown 2",
			"All History Godown 3",
			"Detail Godown",
			"Detail Token Godown",
			"Review Godown",
			"Desk Permission Godown",
			"Other Desk Godown",
			"Seetarambagh Godown",
			"Other Branch Godown",
			"Claimed Role Allowed Godown",
			"Claimed Role Other Godown",
			"Main Location Godown",
			"Restricted Godown",
			"Owner Resolve Godown",
			"Owner Resolve Note Godown",
			"Owner Cancel Godown",
			"Admin Partial Close Godown",
		)

	def tearDown(self):
		frappe.db.rollback()

	def test_customer_can_submit_quantity_only_order_with_financial_year_reference(self):
		product_group = self._create_product_group("Order PG A")
		item = self._create_item("Order Item A", product_group.name)
		customer = self._create_active_customer("9000000201", "ORDER-CUSTOMER-001")

		response = submit_order(
			customer.name,
			[
				{
					"item": item.name,
					"godown": "Main Godown",
					"quantity": 7,
					"stock_shown_at_order_time": 0,
				}
			],
		)

		self.assertTrue(response["success"])
		expected_period = _financial_year_period(frappe.utils.now_datetime())
		self.assertEqual(response["data"]["portal_reference_number"], f"KE-SO-00001-{expected_period}")
		order = frappe.get_doc("Order", response["data"]["order"])
		self.assertEqual(order.status, "Placed")
		self.assertEqual(order.order_source, "Customer")
		self.assertEqual(order.total_item_count, 1)
		self.assertEqual(order.total_quantity, 7)
		self.assertEqual(order.items[0].requested_quantity, 7)
		self.assertEqual(order.godown_allocations[0].stock_shown_at_order_time, 0)

	def test_confirmed_order_lines_cannot_be_edited_after_submission(self):
		product_group = self._create_product_group("Order PG Immutable")
		item = self._create_item("Order Item Immutable", product_group.name)
		customer = self._create_active_customer("9000000431", "ORDER-IMMUTABLE-001")

		response = submit_order(
			customer.name,
			[
				{
					"item": item.name,
					"godown": "Immutable Godown",
					"quantity": 2,
					"stock_shown_at_order_time": 0,
				}
			],
		)
		order = frappe.get_doc("Order", response["data"]["order"])
		order.items[0].requested_quantity = 5
		order.godown_allocations[0].requested_quantity = 5

		with self.assertRaises(frappe.ValidationError):
			order.save(ignore_permissions=True)

	def test_order_submission_rejects_disabled_customer_access(self):
		product_group = self._create_product_group("Order PG Disabled")
		item = self._create_item("Order Item Disabled", product_group.name)
		customer = self._create_active_customer("9000000415", "ORDER-DISABLED-001")
		frappe.db.set_value("Customer", customer.name, "status", "Disabled")
		frappe.db.set_value("Customer", customer.name, "customer_app_access", 0)

		response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Disabled Godown", "quantity": 1}],
		)

		self.assertFalse(response["success"])
		self.assertIn("Customer App Access", response["error"]["message"])

	def test_order_submission_requires_matching_customer_token_when_headers_are_supplied(self):
		product_group = self._create_product_group("Order PG Token")
		item = self._create_item("Order Item Token", product_group.name)
		customer = self._create_active_customer("9000000424", "ORDER-TOKEN-001")
		other_customer = self._create_active_customer("9000000425", "ORDER-TOKEN-002")
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_token_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Token Godown", "quantity": 1}],
			headers={},
		)
		mismatched_token_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Token Godown", "quantity": 1}],
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Token Godown", "quantity": 1}],
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_token_response["success"])
		self.assertIn("token identity", mismatched_token_response["error"]["message"])
		self.assertTrue(valid_response["success"])

	def test_order_submission_revalidates_active_tally_godown_without_requiring_stock(self):
		product_group = self._create_product_group("Order PG Godown Access")
		item = self._create_item("Order Item Godown Access", product_group.name)
		customer = self._create_active_customer("9000000430", "ORDER-GODOWN-001")
		self._create_godown("Active Zero Stock Godown", is_active=1)
		self._create_godown("Inactive Order Godown", is_active=0)

		active_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Active Zero Stock Godown", "quantity": 1}],
		)
		inactive_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Inactive Order Godown", "quantity": 1}],
		)
		unknown_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Unknown Order Godown", "quantity": 1}],
		)

		self.assertTrue(active_response["success"])
		self.assertFalse(inactive_response["success"])
		self.assertFalse(unknown_response["success"])
		self.assertIn("Godown", inactive_response["error"]["message"])
		self.assertIn("Godown", unknown_response["error"]["message"])

	def test_duplicate_item_godown_allocations_are_merged(self):
		product_group = self._create_product_group("Order PG Merge")
		item = self._create_item("Order Item Merge", product_group.name)
		customer = self._create_active_customer("9000000202", "ORDER-CUSTOMER-002")

		response = submit_order(
			customer.name,
			[
				{"item": item.name, "godown": "Main Godown", "quantity": 3},
				{"item": item.name, "godown": "Main Godown", "quantity": 4},
			],
		)

		order = frappe.get_doc("Order", response["data"]["order"])
		self.assertEqual(len(order.items), 1)
		self.assertEqual(order.items[0].requested_quantity, 7)
		self.assertEqual(len(order.godown_allocations), 1)
		self.assertEqual(order.godown_allocations[0].requested_quantity, 7)

	def test_sales_employee_order_stores_internal_note_and_revalidates_access(self):
		shared_group = self._create_product_group("Order PG Shared")
		blocked_group = self._create_product_group("Order PG Blocked")
		allowed_item = self._create_item("Order Item Shared", shared_group.name)
		blocked_item = self._create_item("Order Item Blocked", blocked_group.name)
		customer = self._create_active_customer(
			"9000000203",
			"ORDER-CUSTOMER-003",
			product_groups=[shared_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Order Sales Employee",
				"mobile_number": "9000000204",
				"status": "Active",
				"assigned_customers": [{"customer": customer.name}],
				"product_group_access": [{"product_group": shared_group.name}],
			}
		).insert()
		response = submit_order(
			customer.name,
			[{"item": allowed_item.name, "godown": "Sales Godown", "quantity": 2}],
			sales_employee=sales_employee.name,
			sales_employee_note="Call before dispatch",
		)
		blocked_response = submit_order(
			customer.name,
			[{"item": blocked_item.name, "godown": "Sales Godown", "quantity": 2}],
			sales_employee=sales_employee.name,
		)

		self.assertTrue(response["success"])
		order = frappe.get_doc("Order", response["data"]["order"])
		self.assertEqual(order.order_source, "Sales Employee")
		self.assertEqual(order.sales_employee_note, "Call before dispatch")
		self.assertFalse(blocked_response["success"])

	def test_order_submission_rejects_disabled_sales_employee_access(self):
		product_group = self._create_product_group("Order PG Disabled Sales")
		item = self._create_item("Order Item Disabled Sales", product_group.name)
		customer = self._create_active_customer(
			"9000000416",
			"ORDER-DISABLED-SALES-001",
			product_groups=[product_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Disabled Order Sales Employee",
				"mobile_number": "9000000417",
				"status": "Disabled",
				"assigned_customers": [{"customer": customer.name}],
				"product_group_access": [{"product_group": product_group.name}],
			}
		).insert()

		response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Disabled Sales Godown", "quantity": 2}],
			sales_employee=sales_employee.name,
		)

		self.assertFalse(response["success"])
		self.assertIn("Sales Employee", response["error"]["message"])

	def test_order_submit_creates_customer_pdf_and_whatsapp_log_without_internal_fields(self):
		product_group = self._create_product_group("Order Notification PG")
		item = self._create_item("Order Notification Item", product_group.name)
		customer = self._create_active_customer(
			"9000000215",
			"ORDER-NOTIFY-001",
			product_groups=[product_group.name],
		)
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Notify Sales Employee",
				"mobile_number": "9000000216",
				"status": "Active",
				"assigned_customers": [{"customer": customer.name}],
				"product_group_access": [{"product_group": product_group.name}],
			}
		).insert()

		response = submit_order(
			customer.name,
			[
				{
					"item": item.name,
					"godown": "Notify Godown",
					"quantity": 3,
					"stock_shown_at_order_time": 99,
				}
			],
			sales_employee=sales_employee.name,
			sales_employee_note="Internal note hidden from customer",
		)
		order_name = response["data"]["order"]
		pdf = frappe.get_doc("Order PDF", {"order": order_name})
		notification = frappe.get_doc("Order WhatsApp Notification", {"order": order_name})
		summary_text = pdf.summary_text
		request_payload = json.loads(notification.request_payload)
		request_payload_text = notification.request_payload
		provider_response = json.loads(notification.provider_response)

		self.assertEqual(pdf.customer, customer.name)
		self.assertTrue(pdf.file_url)
		self.assertTrue(pdf.file_url.startswith("/private/files/"))
		self.assertTrue(frappe.db.exists("File", {"file_url": pdf.file_url}))
		self.assertIn("Order Notification Item", summary_text)
		self.assertIn("Notify Godown", summary_text)
		self.assertIn("3", summary_text)
		self.assertIn("Notify Sales Employee", summary_text)
		self.assertNotIn("Internal note hidden from customer", summary_text)
		self.assertNotIn("ORDER-NOTIFY-001", summary_text)
		self.assertNotIn("99", summary_text)
		for forbidden in ("price", "rate", "tax", "value", "amount", "discount"):
			self.assertNotIn(forbidden, summary_text.lower())
		self.assertEqual(notification.recipient_type, "Customer")
		self.assertEqual(notification.recipient_customer, customer.name)
		self.assertFalse(notification.recipient_sales_employee)
		self.assertEqual(notification.order_pdf, pdf.name)
		self.assertEqual(notification.status, "Queued")
		self.assertEqual(request_payload["event"], "Order Placed")
		self.assertEqual(request_payload["order_pdf"], pdf.name)
		self.assertNotIn("Internal note hidden from customer", request_payload_text)
		self.assertNotIn("ORDER-NOTIFY-001", request_payload_text)
		self.assertNotIn("stock_shown_at_order_time", request_payload_text)
		self.assertNotIn("99", request_payload_text)
		for forbidden in ("price", "rate", "tax", "value", "amount", "discount"):
			self.assertNotIn(forbidden, request_payload_text.lower())
		self.assertEqual(provider_response["provider"], "frappe_whatsapp")
		self.assertEqual(provider_response["status"], "Queued")
		self.assertEqual(provider_response["retry_count"], 0)

	def test_order_submission_rejects_non_positive_quantity(self):
		product_group = self._create_product_group("Order PG Quantity")
		item = self._create_item("Order Item Quantity", product_group.name)
		customer = self._create_active_customer("9000000205", "ORDER-CUSTOMER-005")

		response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Main Godown", "quantity": 0}],
		)

		self.assertFalse(response["success"])
		self.assertEqual(response["http_status_code"], 400)

	def test_reference_sequence_resets_by_financial_year(self):
		product_group = self._create_product_group("Order PG Sequence")
		item = self._create_item("Order Item Sequence", product_group.name)
		customer = self._create_active_customer("9000000206", "ORDER-CUSTOMER-006")

		first = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Main Godown", "quantity": 1}],
		)
		second = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Main Godown", "quantity": 1}],
		)

		expected_period = _financial_year_period(frappe.utils.now_datetime())
		self.assertEqual(first["data"]["portal_reference_number"], f"KE-SO-00001-{expected_period}")
		self.assertEqual(second["data"]["portal_reference_number"], f"KE-SO-00002-{expected_period}")

	def test_financial_year_period_uses_indian_april_boundary(self):
		self.assertEqual(_financial_year_period(datetime(2026, 4, 1)), "26-27")
		self.assertEqual(_financial_year_period(datetime(2027, 3, 31)), "26-27")
		self.assertEqual(_financial_year_period(datetime(2027, 4, 1)), "27-28")

	def test_customer_order_history_includes_sales_employee_placed_orders(self):
		product_group = self._create_product_group("Order History PG")
		item = self._create_item("Order History Item", product_group.name)
		customer = self._create_active_customer("9000000207", "ORDER-HISTORY-001")
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "History Sales Employee",
				"mobile_number": "9000000208",
				"status": "Active",
			}
		).insert()
		customer_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Main Godown", "quantity": 1}],
		)
		sales_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Sales Godown", "quantity": 2}],
			sales_employee=sales_employee.name,
			sales_employee_note="Internal only",
		)

		response = order_history(customer.name)

		self.assertTrue(response["success"])
		self.assertEqual(
			{order["name"] for order in response["data"]["orders"]},
			{customer_order["data"]["order"], sales_order["data"]["order"]},
		)

	def test_customer_order_history_requires_matching_customer_token_when_headers_are_supplied(self):
		product_group = self._create_product_group("Order History Token PG")
		item = self._create_item("Order History Token Item", product_group.name)
		customer = self._create_active_customer("9000000426", "ORDER-HISTORY-TOKEN-001")
		other_customer = self._create_active_customer("9000000427", "ORDER-HISTORY-TOKEN-002")
		submit_order(
			customer.name,
			[{"item": item.name, "godown": "History Token Godown", "quantity": 1}],
		)
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_token_response = order_history(customer.name, headers={})
		mismatched_token_response = order_history(
			customer.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_response = order_history(
			customer.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_token_response["success"])
		self.assertIn("token identity", mismatched_token_response["error"]["message"])
		self.assertTrue(valid_response["success"])

	def test_sales_employee_order_history_only_shows_orders_placed_by_employee(self):
		product_group = self._create_product_group("Sales History PG")
		item = self._create_item("Sales History Item", product_group.name)
		customer = self._create_active_customer("9000000209", "ORDER-HISTORY-002")
		first_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "First History Employee",
				"mobile_number": "9000000210",
				"status": "Active",
			}
		).insert()
		second_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Second History Employee",
				"mobile_number": "9000000211",
				"status": "Active",
			}
		).insert()
		visible_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Sales Godown", "quantity": 2}],
			sales_employee=first_employee.name,
		)
		submit_order(
			customer.name,
			[{"item": item.name, "godown": "Other Godown", "quantity": 3}],
			sales_employee=second_employee.name,
		)

		response = order_history(customer.name, sales_employee=first_employee.name)

		self.assertTrue(response["success"])
		self.assertEqual([order["name"] for order in response["data"]["orders"]], [visible_order["data"]["order"]])

	def test_sales_employee_order_history_spans_customers_when_no_customer_filter_is_given(self):
		product_group = self._create_product_group("Sales All History PG")
		item = self._create_item("Sales All History Item", product_group.name)
		first_customer = self._create_active_customer("9000000410", "ORDER-HISTORY-ALL-001")
		second_customer = self._create_active_customer("9000000411", "ORDER-HISTORY-ALL-002")
		other_customer = self._create_active_customer("9000000412", "ORDER-HISTORY-ALL-003")
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "All History Employee",
				"mobile_number": "9000000413",
				"status": "Active",
			}
		).insert()
		other_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Other All History Employee",
				"mobile_number": "9000000414",
				"status": "Active",
			}
		).insert()
		first_order = submit_order(
			first_customer.name,
			[{"item": item.name, "godown": "All History Godown 1", "quantity": 2}],
			sales_employee=sales_employee.name,
		)
		second_order = submit_order(
			second_customer.name,
			[{"item": item.name, "godown": "All History Godown 2", "quantity": 3}],
			sales_employee=sales_employee.name,
		)
		submit_order(
			other_customer.name,
			[{"item": item.name, "godown": "All History Godown 3", "quantity": 4}],
			sales_employee=other_employee.name,
		)

		response = order_history(sales_employee=sales_employee.name)

		self.assertTrue(response["success"])
		self.assertEqual(
			{order["name"] for order in response["data"]["orders"]},
			{first_order["data"]["order"], second_order["data"]["order"]},
		)

	def test_order_detail_hides_internal_fields_from_customer_response(self):
		product_group = self._create_product_group("Order Detail PG")
		item = self._create_item("Order Detail Item", product_group.name)
		customer = self._create_active_customer("9000000212", "ORDER-DETAIL-001")
		sales_employee = frappe.get_doc(
			{
				"doctype": "Sales Employee",
				"sales_employee_name": "Detail Sales Employee",
				"mobile_number": "9000000213",
				"status": "Active",
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[
				{
					"item": item.name,
					"godown": "Detail Godown",
					"quantity": 5,
					"stock_shown_at_order_time": 99,
				}
			],
			sales_employee=sales_employee.name,
			sales_employee_note="Do not show to customer",
		)

		response = order_detail(order_response["data"]["order"], customer=customer.name)

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["placed_by"], "Detail Sales Employee")
		self.assertNotIn("sales_employee_note", response["data"])
		self.assertNotIn("client_code", response["data"])
		self.assertNotIn("stock_shown_at_order_time", response["data"]["godown_allocations"][0])

	def test_customer_order_detail_requires_matching_customer_token_when_headers_are_supplied(self):
		product_group = self._create_product_group("Order Detail Token PG")
		item = self._create_item("Order Detail Token Item", product_group.name)
		customer = self._create_active_customer("9000000428", "ORDER-DETAIL-TOKEN-001")
		other_customer = self._create_active_customer("9000000429", "ORDER-DETAIL-TOKEN-002")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Detail Token Godown", "quantity": 1}],
		)
		valid_token = issue_token("Customer", customer.name)
		other_token = issue_token("Customer", other_customer.name)

		missing_token_response = order_detail(order_response["data"]["order"], customer=customer.name, headers={})
		mismatched_token_response = order_detail(
			order_response["data"]["order"],
			customer=customer.name,
			headers={"Auth-Token": f"Bearer {other_token['access_token']}"},
		)
		valid_response = order_detail(
			order_response["data"]["order"],
			customer=customer.name,
			headers={"Auth-Token": f"Bearer {valid_token['access_token']}"},
		)

		self.assertFalse(missing_token_response["success"])
		self.assertEqual(missing_token_response["http_status_code"], 401)
		self.assertFalse(mismatched_token_response["success"])
		self.assertIn("token identity", mismatched_token_response["error"]["message"])
		self.assertTrue(valid_response["success"])

	def test_manual_review_status_is_under_review_in_mobile_responses(self):
		product_group = self._create_product_group("Order Review PG")
		item = self._create_item("Order Review Item", product_group.name)
		customer = self._create_active_customer("9000000214", "ORDER-REVIEW-001")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Review Godown", "quantity": 1}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Manual Review")

		response = order_detail(order_response["data"]["order"], customer=customer.name)

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["status"], "Manual Review")
		self.assertEqual(response["data"]["display_status"], "Under Review")

	def test_order_permission_conditions_scope_branch_users_by_portal_branch_user_permission(self):
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Desk Permission Branch",
				"is_active": 1,
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Desk Permission Godown",
				"is_active": 1,
			}
		).insert()
		employee = self._create_branch_user("branch.employee@example.com", "Branch Employee", branch.name)
		manager = self._create_branch_user("branch.manager@example.com", "Branch Manager", branch.name)
		product_group = self._create_product_group("Desk Permission PG")
		item = self._create_item("Desk Permission Item", product_group.name)
		customer = self._create_active_customer("9000000901", "DESK-PERM-001")
		visible_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Desk Permission Godown", "quantity": 2}],
		)
		placed_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Desk Permission Godown", "quantity": 1}],
		)
		hidden_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Other Desk Godown", "quantity": 1}],
		)
		frappe.db.set_value("Order", visible_order["data"]["order"], "status", "Completed")
		frappe.db.set_value("Order", hidden_order["data"]["order"], "status", "Placed")
		completed_order = frappe.get_doc("Order", visible_order["data"]["order"])

		employee_condition = order_permission_query(employee.name)
		manager_condition = order_permission_query(manager.name)

		self.assertIn("tabUser Permission", employee_condition)
		self.assertIn("Portal Branch", employee_condition)
		self.assertIn("status", employee_condition)
		self.assertIn("Placed", employee_condition)
		self.assertNotIn("Placed", manager_condition)
		self.assertFalse(order_has_permission(completed_order, user=employee.name))
		self.assertTrue(order_has_permission(completed_order, user=manager.name))

		try:
			frappe.set_user(employee.name)
			employee_orders = {
				row.name
				for row in frappe.get_list("Order", fields=["name"], order_by="name asc")
			}
			frappe.set_user(manager.name)
			manager_orders = {
				row.name
				for row in frappe.get_list("Order", fields=["name"], order_by="name asc")
			}
		finally:
			frappe.set_user("Administrator")

		self.assertIn(placed_order["data"]["order"], employee_orders)
		self.assertNotIn(visible_order["data"]["order"], employee_orders)
		self.assertNotIn(hidden_order["data"]["order"], employee_orders)
		self.assertIn(placed_order["data"]["order"], manager_orders)
		self.assertIn(visible_order["data"]["order"], manager_orders)
		self.assertNotIn(hidden_order["data"]["order"], manager_orders)

	def test_branch_user_with_multiple_branch_permissions_sees_union(self):
		product_group = self._create_product_group("Multi Branch Permission PG")
		item = self._create_item("Multi Branch Permission Item", product_group.name)
		customer = self._create_active_customer("9000000922", "ORDER-MULTI-BRANCH-001")
		branches = []
		for branch_name, godown in (
			("Multi Permission Branch A", "Multi Permission Godown A"),
			("Multi Permission Branch B", "Multi Permission Godown B"),
			("Multi Permission Branch Hidden", "Multi Permission Godown Hidden"),
		):
			branch = frappe.get_doc(
				{
					"doctype": "Portal Branch",
					"branch_name": branch_name,
					"is_active": 1,
				}
			).insert()
			self._create_godown(godown)
			frappe.get_doc(
				{
					"doctype": "Branch Godown Mapping",
					"portal_branch": branch.name,
					"godown": godown,
					"is_active": 1,
				}
			).insert()
			branches.append(branch)
		order_a = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Multi Permission Godown A", "quantity": 1}],
		)
		order_b = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Multi Permission Godown B", "quantity": 1}],
		)
		hidden_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Multi Permission Godown Hidden", "quantity": 1}],
		)
		manager = self._create_branch_user(
			"branch.multi.permission@example.com",
			"Branch Manager",
			branches[0].name,
		)
		frappe.get_doc(
			{
				"doctype": "User Permission",
				"user": manager.name,
				"allow": "Portal Branch",
				"for_value": branches[1].name,
				"apply_to_all_doctypes": 1,
			}
		).insert(ignore_permissions=True)

		try:
			frappe.set_user(manager.name)
			visible_orders = {row.name for row in frappe.get_list("Order", fields=["name"], order_by="name asc")}
		finally:
			frappe.set_user("Administrator")

		self.assertIn(order_a["data"]["order"], visible_orders)
		self.assertIn(order_b["data"]["order"], visible_orders)
		self.assertNotIn(hidden_order["data"]["order"], visible_orders)

	def test_branch_manager_sees_orders_for_mapped_godowns_across_statuses(self):
		product_group = self._create_product_group("Branch Visibility PG")
		item = self._create_item("Branch Visibility Item", product_group.name)
		customer = self._create_active_customer("9000000215", "ORDER-BRANCH-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Seetarambagh",
				"is_active": 1,
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Seetarambagh Godown",
				"is_active": 1,
			}
		).insert()
		visible_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Seetarambagh Godown", "quantity": 2}],
		)
		hidden_order = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Other Branch Godown", "quantity": 2}],
		)
		frappe.db.set_value("Order", visible_order["data"]["order"], "status", "Manual Review")
		frappe.db.set_value("Order", hidden_order["data"]["order"], "status", "Manual Review")
		frappe.get_doc(
			{
				"doctype": "Order Reconciliation Log",
				"order": visible_order["data"]["order"],
				"status": "Manual Review",
				"reason_code": "CUSTOMER_CLIENT_CODE_MISMATCH",
				"message": "Customer Client Code mismatch for branch-visible order",
				"created_at": "2026-05-19 12:00:00",
			}
		).insert(ignore_permissions=True)

		response = branch_visible_orders(branch.name, role="Branch Manager")

		self.assertTrue(response["success"])
		self.assertEqual([order["name"] for order in response["data"]["orders"]], [visible_order["data"]["order"]])
		self.assertEqual(response["data"]["orders"][0]["status"], "Manual Review")
		self.assertEqual(response["data"]["orders"][0]["display_status"], "Under Review")
		self.assertEqual(
			response["data"]["orders"][0]["manual_review_reason"],
			"Customer Client Code mismatch for branch-visible order",
		)
		self.assertEqual(
			response["data"]["orders"][0]["manual_review_reason_code"],
			"CUSTOMER_CLIENT_CODE_MISMATCH",
		)

	def test_mixed_branch_manager_employee_visibility_uses_manager_scope(self):
		product_group = self._create_product_group("Mixed Branch Role PG")
		item = self._create_item("Mixed Branch Role Item", product_group.name)
		customer = self._create_active_customer("9000000916", "ORDER-BRANCH-MIXED-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Mixed Role Branch",
				"is_active": 1,
			}
		).insert()
		self._create_godown("Mixed Role Godown")
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Mixed Role Godown",
				"is_active": 1,
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Mixed Role Godown", "quantity": 2}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Completed")
		mixed_user = self._create_branch_user(
			"branch.mixed.role@example.com",
			"Branch Manager",
			branch.name,
		)
		mixed_user.add_roles("Branch Employee")

		try:
			frappe.set_user(mixed_user.name)
			response = branch_visible_orders(branch.name, role="Branch Employee")
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(response["success"])
		self.assertEqual([order["name"] for order in response["data"]["orders"]], [order_response["data"]["order"]])

	def test_branch_manager_can_move_visible_placed_order_to_processing_with_status_log(self):
		product_group = self._create_product_group("Branch Manager Processing PG")
		item = self._create_item("Branch Manager Processing Item", product_group.name)
		customer = self._create_active_customer("9000000917", "ORDER-BRANCH-MANAGER-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Manager Processing Branch",
				"is_active": 1,
			}
		).insert()
		self._create_godown("Manager Processing Godown")
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Manager Processing Godown",
				"is_active": 1,
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Manager Processing Godown", "quantity": 2}],
		)
		manager = self._create_branch_user(
			"branch.manager.processing@example.com",
			"Branch Manager",
			branch.name,
		)

		try:
			frappe.set_user(manager.name)
			response = branch_mark_processing(branch.name, order_response["data"]["order"], role="Branch Manager")
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(response["success"])
		order = frappe.get_doc("Order", order_response["data"]["order"])
		status_log = frappe.get_last_doc("Order Status Log", filters={"order": order.name})
		self.assertEqual(order.status, "Processing")
		self.assertEqual(status_log.from_status, "Placed")
		self.assertEqual(status_log.to_status, "Processing")
		self.assertEqual(status_log.role, "Branch Manager")

	def test_order_form_exposes_branch_move_to_processing_action(self):
		script = (
			Path(__file__).parents[1]
			/ "kunal_enterprises"
			/ "doctype"
			/ "order"
			/ "order.js"
		).read_text()

		self.assertIn('frm.doc.status === "Placed"', script)
		self.assertIn('__("Move to Processing")', script)
		self.assertIn("kunal_enterprises.api.branch_orders.mark_visible_order_processing", script)

	def test_branch_desk_action_moves_visible_placed_order_to_processing(self):
		product_group = self._create_product_group("Branch Desk Processing PG")
		item = self._create_item("Branch Desk Processing Item", product_group.name)
		customer = self._create_active_customer("9000000926", "ORDER-BRANCH-DESK-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Desk Processing Branch",
				"is_active": 1,
			}
		).insert()
		self._create_godown("Desk Processing Godown")
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Desk Processing Godown",
				"is_active": 1,
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Desk Processing Godown", "quantity": 2}],
		)
		employee = self._create_branch_user(
			"branch.desk.processing@example.com",
			"Branch Employee",
			branch.name,
		)
		process_order = frappe.get_attr(
			"kunal_enterprises.api.branch_orders.mark_visible_order_processing"
		)

		try:
			frappe.set_user(employee.name)
			response = process_order(order_response["data"]["order"])
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(response["success"])
		order = frappe.get_doc("Order", order_response["data"]["order"])
		status_log = frappe.get_last_doc("Order Status Log", filters={"order": order.name})
		self.assertEqual(order.status, "Processing")
		self.assertEqual(status_log.from_status, "Placed")
		self.assertEqual(status_log.to_status, "Processing")
		self.assertEqual(status_log.role, "Branch Employee")

	def test_branch_desk_action_rejects_order_outside_user_branch(self):
		product_group = self._create_product_group("Branch Desk Hidden PG")
		item = self._create_item("Branch Desk Hidden Item", product_group.name)
		customer = self._create_active_customer("9000000927", "ORDER-BRANCH-DESK-002")
		allowed_branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Desk Allowed Branch",
				"is_active": 1,
			}
		).insert()
		hidden_branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Desk Hidden Branch",
				"is_active": 1,
			}
		).insert()
		for branch, godown in (
			(allowed_branch.name, "Desk Allowed Godown"),
			(hidden_branch.name, "Desk Hidden Godown"),
		):
			self._create_godown(godown)
			frappe.get_doc(
				{
					"doctype": "Branch Godown Mapping",
					"portal_branch": branch,
					"godown": godown,
					"is_active": 1,
				}
			).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Desk Hidden Godown", "quantity": 2}],
		)
		employee = self._create_branch_user(
			"branch.desk.hidden@example.com",
			"Branch Employee",
			allowed_branch.name,
		)
		process_order = frappe.get_attr(
			"kunal_enterprises.api.branch_orders.mark_visible_order_processing"
		)

		try:
			frappe.set_user(employee.name)
			response = process_order(order_response["data"]["order"])
		finally:
			frappe.set_user("Administrator")

		order = frappe.get_doc("Order", order_response["data"]["order"])
		self.assertFalse(response["success"])
		self.assertEqual(order.status, "Placed")
		self.assertIn("not visible", response["error"]["message"])

	def test_branch_order_apis_do_not_trust_claimed_role_or_branch(self):
		product_group = self._create_product_group("Branch Claimed Role PG")
		item = self._create_item("Branch Claimed Role Item", product_group.name)
		customer = self._create_active_customer("9000000915", "ORDER-BRANCH-CLAIM-001")
		allowed_branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Claimed Role Allowed Branch",
				"is_active": 1,
			}
		).insert()
		other_branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Claimed Role Other Branch",
				"is_active": 1,
			}
		).insert()
		for branch, godown in (
			(allowed_branch.name, "Claimed Role Allowed Godown"),
			(other_branch.name, "Claimed Role Other Godown"),
		):
			frappe.get_doc(
				{
					"doctype": "Branch Godown Mapping",
					"portal_branch": branch,
					"godown": godown,
					"is_active": 1,
				}
			).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Claimed Role Allowed Godown", "quantity": 2}],
		)
		branch_user = self._create_branch_user(
			"branch.claimed.role@example.com",
			"Branch Employee",
			allowed_branch.name,
		)
		plain_user = self._create_role_user("branch.plain@example.com", "Admin")

		try:
			frappe.set_user(plain_user.name)
			forged_visible_response = branch_visible_orders(allowed_branch.name, role="Branch Manager")
			frappe.set_user(branch_user.name)
			wrong_branch_response = branch_visible_orders(other_branch.name, role="Branch Employee")
			forged_manager_response = branch_visible_orders(allowed_branch.name, role="Branch Manager")
			processing_response = branch_mark_processing(
				allowed_branch.name,
				order_response["data"]["order"],
				role="Branch Employee",
			)
		finally:
			frappe.set_user("Administrator")

		self.assertFalse(forged_visible_response["success"])
		self.assertFalse(wrong_branch_response["success"])
		self.assertFalse(forged_manager_response["success"])
		self.assertTrue(processing_response["success"])

	def test_branch_employee_can_move_visible_placed_order_to_processing(self):
		product_group = self._create_product_group("Branch Employee PG")
		item = self._create_item("Branch Employee Item", product_group.name)
		customer = self._create_active_customer("9000000216", "ORDER-BRANCH-002")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Main Location",
				"is_active": 1,
			}
		).insert()
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Main Location Godown",
				"is_active": 1,
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Main Location Godown", "quantity": 2}],
		)

		transition_response = branch_mark_processing(branch.name, order_response["data"]["order"], role="Branch Employee")
		visible_response = branch_visible_orders(branch.name, role="Branch Employee")
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertTrue(transition_response["success"])
		self.assertEqual(order.status, "Processing")
		self.assertEqual([row["name"] for row in visible_response["data"]["orders"]], [order.name])
		self.assertEqual(visible_response["data"]["orders"][0]["status"], "Processing")
		status_log = frappe.get_last_doc("Order Status Log", filters={"order": order.name})
		self.assertEqual(status_log.from_status, "Placed")
		self.assertEqual(status_log.to_status, "Processing")
		self.assertEqual(status_log.role, "Branch Employee")

	def test_owner_and_admin_can_move_placed_order_to_processing_with_status_log(self):
		process_order = getattr(owner_admin_order_controls, "mark_processing", None)
		self.assertIsNotNone(process_order)

		for role, email, mobile, client_code in (
			("Owner", "owner.processing@example.com", "9000000918", "ORDER-OWNER-PROCESS-001"),
			("Admin", "admin.processing@example.com", "9000000919", "ORDER-ADMIN-PROCESS-001"),
		):
			with self.subTest(role=role):
				product_group = self._create_product_group(f"{role} Processing PG")
				item = self._create_item(f"{role} Processing Item", product_group.name)
				customer = self._create_active_customer(mobile, client_code)
				self._create_godown(f"{role} Processing Godown")
				order_response = submit_order(
					customer.name,
					[{"item": item.name, "godown": f"{role} Processing Godown", "quantity": 2}],
				)
				user = self._create_role_user(email, role)

				try:
					frappe.set_user(user.name)
					response = process_order(order_response["data"]["order"], role=role)
				finally:
					frappe.set_user("Administrator")

				order = frappe.get_doc("Order", order_response["data"]["order"])
				status_log = frappe.get_last_doc("Order Status Log", filters={"order": order.name})
				self.assertTrue(response["success"])
				self.assertEqual(order.status, "Processing")
				self.assertEqual(status_log.from_status, "Placed")
				self.assertEqual(status_log.to_status, "Processing")
				self.assertEqual(status_log.role, role)

	def test_processing_actions_reject_non_placed_orders_without_status_log(self):
		process_order = getattr(owner_admin_order_controls, "mark_processing", None)
		self.assertIsNotNone(process_order)
		product_group = self._create_product_group("Processing Reject PG")
		item = self._create_item("Processing Reject Item", product_group.name)
		customer = self._create_active_customer("9000000920", "ORDER-PROCESS-REJECT-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Processing Reject Branch",
				"is_active": 1,
			}
		).insert()
		self._create_godown("Processing Reject Godown")
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Processing Reject Godown",
				"is_active": 1,
			}
		).insert()
		branch_order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Processing Reject Godown", "quantity": 2}],
		)
		owner_order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Processing Reject Godown", "quantity": 1}],
		)
		frappe.db.set_value("Order", branch_order_response["data"]["order"], "status", "Processing")
		frappe.db.set_value("Order", owner_order_response["data"]["order"], "status", "Processing")
		branch_user = self._create_branch_user(
			"branch.processing.reject@example.com",
			"Branch Employee",
			branch.name,
		)
		owner_user = self._create_role_user("owner.processing.reject@example.com", "Owner")

		try:
			frappe.set_user(branch_user.name)
			branch_response = branch_mark_processing(
				branch.name,
				branch_order_response["data"]["order"],
				role="Branch Employee",
			)
			frappe.set_user(owner_user.name)
			owner_response = process_order(owner_order_response["data"]["order"], role="Owner")
		finally:
			frappe.set_user("Administrator")

		self.assertFalse(branch_response["success"])
		self.assertFalse(owner_response["success"])
		self.assertEqual(
			frappe.db.count("Order Status Log", {"order": branch_order_response["data"]["order"]}),
			0,
		)
		self.assertEqual(
			frappe.db.count("Order Status Log", {"order": owner_order_response["data"]["order"]}),
			0,
		)

	def test_branch_processing_rejects_order_without_mapped_godown(self):
		product_group = self._create_product_group("Unmapped Branch Processing PG")
		item = self._create_item("Unmapped Branch Processing Item", product_group.name)
		customer = self._create_active_customer("9000000921", "ORDER-BRANCH-UNMAPPED-001")
		branch = frappe.get_doc(
			{
				"doctype": "Portal Branch",
				"branch_name": "Unmapped Processing Branch",
				"is_active": 1,
			}
		).insert()
		self._create_godown("Mapped Processing Godown")
		self._create_godown("Unmapped Processing Godown")
		frappe.get_doc(
			{
				"doctype": "Branch Godown Mapping",
				"portal_branch": branch.name,
				"godown": "Mapped Processing Godown",
				"is_active": 1,
			}
		).insert()
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Unmapped Processing Godown", "quantity": 2}],
		)
		branch_user = self._create_branch_user(
			"branch.processing.unmapped@example.com",
			"Branch Employee",
			branch.name,
		)

		try:
			frappe.set_user(branch_user.name)
			response = branch_mark_processing(branch.name, order_response["data"]["order"], role="Branch Employee")
		finally:
			frappe.set_user("Administrator")

		order = frappe.get_doc("Order", order_response["data"]["order"])
		self.assertFalse(response["success"])
		self.assertEqual(order.status, "Placed")
		self.assertEqual(frappe.db.count("Order Status Log", {"order": order.name}), 0)

	def test_branch_roles_cannot_cancel_partially_close_or_resolve_manual_review(self):
		product_group = self._create_product_group("Branch Restricted Control PG")
		item = self._create_item("Branch Restricted Control Item", product_group.name)
		customer = self._create_active_customer("9000000217", "ORDER-BRANCH-003")
		branch_user = self._create_role_user("branch.control@example.com", "Branch Manager")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Restricted Godown", "quantity": 2}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Manual Review")

		cancel_response = cancel_order(order_response["data"]["order"], role="Branch Manager", note="No branch cancel")
		partial_response = partially_close_order(
			order_response["data"]["order"],
			role="Branch Employee",
			note="No branch partial close",
		)
		resolve_response = resolve_manual_review(
			order_response["data"]["order"],
			role="Branch Manager",
			resolution_note="No branch resolution",
		)
		try:
			frappe.set_user(branch_user.name)
			forged_owner_response = cancel_order(
				order_response["data"]["order"],
				role="Owner",
				note="Forged owner claim",
			)
		finally:
			frappe.set_user("Administrator")
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertFalse(cancel_response["success"])
		self.assertFalse(partial_response["success"])
		self.assertFalse(resolve_response["success"])
		self.assertFalse(forged_owner_response["success"])
		self.assertEqual(order.status, "Manual Review")

	def test_owner_can_resolve_manual_review_with_note(self):
		product_group = self._create_product_group("Owner Resolve PG")
		item = self._create_item("Owner Resolve Item", product_group.name)
		customer = self._create_active_customer("9000000218", "ORDER-OWNER-001")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Owner Resolve Godown", "quantity": 2}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Manual Review")

		response = resolve_manual_review(
			order_response["data"]["order"],
			role="Owner",
			resolution_note="Verified in Tally and resumed processing",
		)
		order = frappe.get_doc("Order", order_response["data"]["order"])
		status_log = frappe.get_doc("Order Status Log", {"order": order.name, "to_status": "Processing"})

		self.assertTrue(response["success"])
		self.assertEqual(order.status, "Processing")
		self.assertEqual(status_log.from_status, "Manual Review")
		self.assertEqual(status_log.role, "Owner")
		self.assertEqual(status_log.note, "Verified in Tally and resumed processing")

	def test_manual_review_resolution_requires_note(self):
		product_group = self._create_product_group("Owner Resolve Note PG")
		item = self._create_item("Owner Resolve Note Item", product_group.name)
		customer = self._create_active_customer("9000000431", "ORDER-OWNER-NOTE-001")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Owner Resolve Note Godown", "quantity": 2}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Manual Review")

		response = resolve_manual_review(
			order_response["data"]["order"],
			role="Admin",
			resolution_note="",
		)
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertFalse(response["success"])
		self.assertIn("Resolution note", response["error"]["message"])
		self.assertEqual(order.status, "Manual Review")

	def test_owner_can_cancel_order_with_status_log(self):
		product_group = self._create_product_group("Owner Cancel PG")
		item = self._create_item("Owner Cancel Item", product_group.name)
		customer = self._create_active_customer("9000000219", "ORDER-OWNER-002")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Owner Cancel Godown", "quantity": 2}],
		)

		response = cancel_order(
			order_response["data"]["order"],
			role="Owner",
			note="Customer requested cancellation",
		)
		order = frappe.get_doc("Order", order_response["data"]["order"])
		status_log = frappe.get_doc("Order Status Log", {"order": order.name, "to_status": "Cancelled"})

		self.assertTrue(response["success"])
		self.assertEqual(order.status, "Cancelled")
		self.assertEqual(response["data"]["status"], "Cancelled")
		self.assertEqual(status_log.from_status, "Placed")
		self.assertEqual(status_log.role, "Owner")
		self.assertEqual(status_log.note, "Customer requested cancellation")

	def test_admin_can_partially_close_order_with_status_log(self):
		product_group = self._create_product_group("Admin Partial Close PG")
		item = self._create_item("Admin Partial Close Item", product_group.name)
		customer = self._create_active_customer("9000000220", "ORDER-ADMIN-001")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Admin Partial Close Godown", "quantity": 4}],
		)
		frappe.db.set_value("Order", order_response["data"]["order"], "status", "Partially Processed")

		response = partially_close_order(
			order_response["data"]["order"],
			role="Admin",
			note="Remaining quantity will not be supplied",
		)
		order = frappe.get_doc("Order", order_response["data"]["order"])
		status_log = frappe.get_doc("Order Status Log", {"order": order.name, "to_status": "Partially Closed"})

		self.assertTrue(response["success"])
		self.assertEqual(order.status, "Partially Closed")
		self.assertEqual(response["data"]["status"], "Partially Closed")
		self.assertEqual(status_log.from_status, "Partially Processed")
		self.assertEqual(status_log.role, "Admin")
		self.assertEqual(status_log.note, "Remaining quantity will not be supplied")

	def _create_product_group(self, group_name):
		return frappe.get_doc(
			{
				"doctype": "Tally Stock Group",
				"group_name": group_name,
				"is_root": 1,
				"depth": 0,
				"full_path": group_name,
				"is_active": 1,
			}
		).insert()

	def _create_item(self, item_name, root_stock_group):
		return frappe.get_doc(
			{
				"doctype": "Tally Item",
				"item_name": item_name,
				"root_stock_group": root_stock_group,
				"uom": "PCS",
				"total_closing_balance": 0,
				"is_active": 1,
			}
		).insert()

	def _create_godown(self, godown_name, is_active=1):
		if frappe.db.exists("Tally Godown", godown_name):
			frappe.db.set_value("Tally Godown", godown_name, "is_active", is_active)
			return frappe.get_doc("Tally Godown", godown_name)
		return frappe.get_doc(
			{
				"doctype": "Tally Godown",
				"godown_name": godown_name,
				"is_active": is_active,
			}
		).insert()

	def _ensure_godowns(self, *godown_names):
		for godown_name in godown_names:
			self._create_godown(godown_name)

	def _create_role_user(self, email, role):
		user = frappe.get_doc(
			{
				"doctype": "User",
				"email": email,
				"first_name": email.split("@")[0],
				"enabled": 1,
				"send_welcome_email": 0,
			}
		).insert(ignore_permissions=True)
		user.add_roles(role)
		return user

	def _create_branch_user(self, email, role, branch):
		user = self._create_role_user(email, role)
		frappe.get_doc(
			{
				"doctype": "User Permission",
				"user": user.name,
				"allow": "Portal Branch",
				"for_value": branch,
				"apply_to_all_doctypes": 1,
			}
		).insert(ignore_permissions=True)
		return user

	def _create_active_customer(self, mobile_number, client_code, product_groups=None):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": client_code,
				"ledger_name": f"Ledger {client_code}",
				"is_active": 1,
			}
		).insert()
		return frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": f"Customer {client_code}",
				"business_legal_name": f"Business {client_code}",
				"mobile_number": mobile_number,
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": client_code,
				"product_group_access": [
					{"product_group": product_group}
					for product_group in (product_groups or [])
				],
			}
		).insert()


class TestOrderReconciliation(FrappeTestCase):
	def setUp(self):
		self._create_godown("Recon Godown")

	def tearDown(self):
		frappe.db.rollback()

	def test_sales_invoice_partially_fulfills_order_item(self):
		product_group = self._create_product_group("Recon PG")
		item = self._create_item("Recon Item", product_group.name)
		customer = self._create_active_customer("9000000301", "RECON-CUSTOMER-001")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 5}],
		)
		frappe.get_doc(
			{
				"doctype": "Tally Voucher",
				"voucher_type": "Sales Invoice",
				"voucher_number": "SI-RECON-001",
				"reference_number": order_response["data"]["portal_reference_number"],
				"party_client_code": "RECON-CUSTOMER-001",
				"tracking_number": "TRACK-RECON-001",
				"voucher_date": "2026-05-19",
				"lines": [
					{
						"item": item.name,
						"godown": "Recon Godown",
						"quantity": 2,
						"tracking_number": "TRACK-RECON-001",
					}
				],
			}
		).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Partially Processed")
		self.assertEqual(order.items[0].fulfilled_quantity, 2)
		self.assertEqual(order.items[0].pending_quantity, 3)
		self.assertEqual(order.items[0].status, "Partially Processed")

	def test_sales_invoice_completes_order_when_all_items_are_fulfilled(self):
		product_group = self._create_product_group("Recon Complete PG")
		item_a = self._create_item("Recon Complete Item A", product_group.name)
		item_b = self._create_item("Recon Complete Item B", product_group.name)
		customer = self._create_active_customer("9000000302", "RECON-CUSTOMER-002")
		order_response = submit_order(
			customer.name,
			[
				{"item": item_a.name, "godown": "Recon Godown", "quantity": 2},
				{"item": item_b.name, "godown": "Recon Godown", "quantity": 3},
			],
		)
		frappe.get_doc(
			{
				"doctype": "Tally Voucher",
				"voucher_type": "Sales Invoice",
				"voucher_number": "SI-RECON-002",
				"reference_number": order_response["data"]["portal_reference_number"],
				"party_client_code": "RECON-CUSTOMER-002",
				"tracking_number": "TRACK-RECON-002",
				"voucher_date": "2026-05-19",
				"lines": [
					{
						"item": item_a.name,
						"godown": "Recon Godown",
						"quantity": 2,
						"tracking_number": "TRACK-RECON-002",
					},
					{
						"item": item_b.name,
						"godown": "Recon Godown",
						"quantity": 3,
						"tracking_number": "TRACK-RECON-002",
					},
				],
			}
		).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Completed")
		self.assertEqual({row.item: row.fulfilled_quantity for row in order.items}, {item_a.name: 2, item_b.name: 3})
		self.assertEqual({row.item: row.pending_quantity for row in order.items}, {item_a.name: 0, item_b.name: 0})
		self.assertEqual({row.status for row in order.items}, {"Completed"})

	def test_multiple_sales_invoices_cumulatively_complete_order(self):
		product_group = self._create_product_group("Recon Cumulative PG")
		item = self._create_item("Recon Cumulative Item", product_group.name)
		customer = self._create_active_customer("9000000306", "RECON-CUSTOMER-006")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 5}],
		)
		for voucher_number, quantity, tracking_number in (
			("SI-RECON-006-A", 2, "TRACK-RECON-006-A"),
			("SI-RECON-006-B", 3, "TRACK-RECON-006-B"),
		):
			frappe.get_doc(
				{
					"doctype": "Tally Voucher",
					"voucher_type": "Sales Invoice",
					"voucher_number": voucher_number,
					"reference_number": order_response["data"]["portal_reference_number"],
					"party_client_code": "RECON-CUSTOMER-006",
					"tracking_number": tracking_number,
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Recon Godown",
							"quantity": quantity,
							"tracking_number": tracking_number,
						}
					],
				}
			).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Completed")
		self.assertEqual(order.items[0].fulfilled_quantity, 5)
		self.assertEqual(order.items[0].pending_quantity, 0)
		self.assertEqual(order.items[0].status, "Completed")

	def test_delivery_challan_is_not_double_counted_when_matching_sales_invoice_exists(self):
		product_group = self._create_product_group("Recon Mirror PG")
		item = self._create_item("Recon Mirror Item", product_group.name)
		customer = self._create_active_customer("9000000307", "RECON-CUSTOMER-007")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 3}],
		)
		for voucher_type, voucher_number in (
			("Delivery Challan", "DC-RECON-007"),
			("Sales Invoice", "SI-RECON-007"),
		):
			frappe.get_doc(
				{
					"doctype": "Tally Voucher",
					"voucher_type": voucher_type,
					"voucher_number": voucher_number,
					"reference_number": order_response["data"]["portal_reference_number"],
					"party_client_code": "RECON-CUSTOMER-007",
					"tracking_number": "TRACK-RECON-007",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Recon Godown",
							"quantity": 3,
							"tracking_number": "TRACK-RECON-007",
						}
					],
				}
			).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])
		logs = frappe.get_all(
			"Order Reconciliation Log",
			filters={"order": order.name},
			fields=["status", "message"],
			order_by="creation asc",
		)

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Completed")
		self.assertEqual(order.items[0].fulfilled_quantity, 3)
		self.assertEqual(order.items[0].pending_quantity, 0)
		self.assertNotIn("Manual Review", {log.status for log in logs})
		self.assertTrue(any("Delivery Challan superseded by Sales Invoice" in log.message for log in logs))

	def test_ambiguous_duplicate_sales_invoice_movement_moves_order_to_manual_review(self):
		product_group = self._create_product_group("Recon Ambiguous PG")
		item = self._create_item("Recon Ambiguous Item", product_group.name)
		customer = self._create_active_customer("9000000308", "RECON-CUSTOMER-008")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 2}],
		)
		for voucher_number in ("SI-RECON-008-A", "SI-RECON-008-B"):
			frappe.get_doc(
				{
					"doctype": "Tally Voucher",
					"voucher_type": "Sales Invoice",
					"voucher_number": voucher_number,
					"reference_number": order_response["data"]["portal_reference_number"],
					"party_client_code": "RECON-CUSTOMER-008",
					"tracking_number": "TRACK-RECON-008",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Recon Godown",
							"quantity": 1,
							"tracking_number": "TRACK-RECON-008",
						}
					],
				}
			).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])
		logs = frappe.get_all(
			"Order Reconciliation Log",
			filters={"order": order.name},
			fields=["status", "message"],
			order_by="creation asc",
		)

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Manual Review")
		self.assertTrue(any(log.status == "Manual Review" for log in logs))
		self.assertTrue(any("Ambiguous duplicate movement" in log.message for log in logs))

	def test_over_fulfillment_moves_order_to_manual_review_with_reason(self):
		product_group = self._create_product_group("Recon Over PG")
		item = self._create_item("Recon Over Item", product_group.name)
		customer = self._create_active_customer("9000000303", "RECON-CUSTOMER-003")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 4}],
		)
		frappe.get_doc(
			{
				"doctype": "Tally Voucher",
				"voucher_type": "Sales Invoice",
				"voucher_number": "SI-RECON-003",
				"reference_number": order_response["data"]["portal_reference_number"],
				"party_client_code": "RECON-CUSTOMER-003",
				"tracking_number": "TRACK-RECON-003",
				"voucher_date": "2026-05-19",
				"lines": [
					{
						"item": item.name,
						"godown": "Recon Godown",
						"quantity": 5,
						"tracking_number": "TRACK-RECON-003",
					}
				],
			}
		).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])
		log = frappe.get_doc("Order Reconciliation Log", {"order": order.name})

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Manual Review")
		self.assertEqual(log.status, "Manual Review")
		self.assertEqual(log.reason_code, "OVER_FULFILLMENT")
		self.assertIn("Over fulfillment", log.message)

	def test_extra_voucher_item_moves_order_to_manual_review_with_reason(self):
		product_group = self._create_product_group("Recon Extra PG")
		ordered_item = self._create_item("Recon Ordered Item", product_group.name)
		extra_item = self._create_item("Recon Extra Item", product_group.name)
		customer = self._create_active_customer("9000000304", "RECON-CUSTOMER-004")
		order_response = submit_order(
			customer.name,
			[{"item": ordered_item.name, "godown": "Recon Godown", "quantity": 4}],
		)
		frappe.get_doc(
			{
				"doctype": "Tally Voucher",
				"voucher_type": "Sales Invoice",
				"voucher_number": "SI-RECON-004",
				"reference_number": order_response["data"]["portal_reference_number"],
				"party_client_code": "RECON-CUSTOMER-004",
				"tracking_number": "TRACK-RECON-004",
				"voucher_date": "2026-05-19",
				"lines": [
					{
						"item": ordered_item.name,
						"godown": "Recon Godown",
						"quantity": 4,
						"tracking_number": "TRACK-RECON-004",
					},
					{
						"item": extra_item.name,
						"godown": "Recon Godown",
						"quantity": 1,
						"tracking_number": "TRACK-RECON-004",
					},
				],
			}
		).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])
		log = frappe.get_doc("Order Reconciliation Log", {"order": order.name})

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Manual Review")
		self.assertEqual(log.status, "Manual Review")
		self.assertIn("Extra unmatched item", log.message)
		self.assertIn(extra_item.name, log.message)

	def test_customer_mismatch_moves_order_to_manual_review_with_reason_context(self):
		product_group = self._create_product_group("Recon Customer Mismatch PG")
		item = self._create_item("Recon Customer Mismatch Item", product_group.name)
		customer = self._create_active_customer("9000000305", "RECON-CUSTOMER-005")
		order_response = submit_order(
			customer.name,
			[{"item": item.name, "godown": "Recon Godown", "quantity": 2}],
		)
		frappe.get_doc(
			{
				"doctype": "Tally Voucher",
				"voucher_type": "Sales Invoice",
				"voucher_number": "SI-RECON-005",
				"reference_number": order_response["data"]["portal_reference_number"],
				"party_client_code": "DIFFERENT-CUSTOMER",
				"tracking_number": "TRACK-RECON-005",
				"voucher_date": "2026-05-19",
				"lines": [
					{
						"item": item.name,
						"godown": "Recon Godown",
						"quantity": 2,
						"tracking_number": "TRACK-RECON-005",
					}
				],
			}
		).insert(ignore_permissions=True)

		run = run_reconciliation()
		order = frappe.get_doc("Order", order_response["data"]["order"])
		log = frappe.get_doc("Order Reconciliation Log", {"order": order.name})

		self.assertEqual(run.status, "Completed")
		self.assertEqual(order.status, "Manual Review")
		self.assertEqual(log.status, "Manual Review")
		self.assertIn("Customer Client Code mismatch", log.message)
		self.assertIn("RECON-CUSTOMER-005", log.message)
		self.assertIn("DIFFERENT-CUSTOMER", log.message)

	def _create_product_group(self, group_name):
		return frappe.get_doc(
			{
				"doctype": "Tally Stock Group",
				"group_name": group_name,
				"is_root": 1,
				"depth": 0,
				"full_path": group_name,
				"is_active": 1,
			}
		).insert()

	def _create_item(self, item_name, root_stock_group):
		return frappe.get_doc(
			{
				"doctype": "Tally Item",
				"item_name": item_name,
				"root_stock_group": root_stock_group,
				"uom": "PCS",
				"total_closing_balance": 0,
				"is_active": 1,
			}
		).insert()

	def _create_godown(self, godown_name, is_active=1):
		if frappe.db.exists("Tally Godown", godown_name):
			frappe.db.set_value("Tally Godown", godown_name, "is_active", is_active)
			return frappe.get_doc("Tally Godown", godown_name)
		return frappe.get_doc(
			{
				"doctype": "Tally Godown",
				"godown_name": godown_name,
				"is_active": is_active,
			}
		).insert()

	def _create_active_customer(self, mobile_number, client_code):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": client_code,
				"ledger_name": f"Ledger {client_code}",
				"is_active": 1,
			}
		).insert()
		return frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": f"Customer {client_code}",
				"business_legal_name": f"Business {client_code}",
				"mobile_number": mobile_number,
				"mobile_verified": 1,
				"admin_approved": 1,
				"status": "Active",
				"client_code": client_code,
			}
		).insert()

class TestTallyStockSync(FrappeTestCase):
	def setUp(self):
		self._create_godown("Manual Sync Godown")

	def tearDown(self):
		frappe.db.rollback()

	def test_tally_master_sync_imports_units_godowns_categories_and_records_run(self):
		run = sync_tally_masters(
			{
				"units": [{"unit_name": "BOX", "tally_guid": "UOM-SYNC-BOX", "symbol": "BX", "is_active": 1}],
				"godowns": [{"godown_name": "Sync Main Godown", "tally_guid": "GODOWN-SYNC-MAIN", "is_active": 1}],
				"stock_categories": [{"category_name": "Sync Category", "tally_guid": "CATEGORY-SYNC-001", "is_active": 1}],
			}
		)

		self.assertEqual(run.status, "Completed")
		self.assertEqual(run.records_seen, 3)
		self.assertEqual(run.records_processed, 3)
		self.assertTrue(frappe.db.exists("Tally Unit", "BOX"))
		self.assertTrue(frappe.db.exists("Tally Godown", "Sync Main Godown"))
		self.assertTrue(frappe.db.exists("Tally Stock Category", "Sync Category"))

	def test_tally_master_sync_imports_stock_groups_items_and_customer_ledgers(self):
		run = sync_tally_masters(
			{
				"stock_groups": [
					{
						"group_name": "Sync Root Group",
						"tally_guid": "SG-SYNC-ROOT",
						"is_root": 1,
						"depth": 0,
						"full_path": "Sync Root Group",
						"is_active": 1,
					},
					{
						"group_name": "Sync Child Group",
						"tally_guid": "SG-SYNC-CHILD",
						"parent_stock_group": "Sync Root Group",
						"root_stock_group": "Sync Root Group",
						"depth": 1,
						"full_path": "Sync Root Group > Sync Child Group",
						"is_active": 1,
					},
				],
				"items": [
					{
						"item_name": "Sync Imported Item",
						"tally_guid": "ITEM-SYNC-001",
						"immediate_stock_group": "Sync Child Group",
						"root_stock_group": "Sync Root Group",
						"stock_category": "Sync Category",
						"uom": "PCS",
						"total_closing_balance": 12,
						"is_active": 1,
					}
				],
				"customer_ledgers": [
					{
						"client_code": "SYNC-LEDGER-001",
						"ledger_name": "Sync Ledger Customer",
						"tally_guid": "LEDGER-SYNC-001",
						"is_active": 1,
					}
				],
			}
		)
		root_group = frappe.get_doc("Tally Stock Group", "Sync Root Group")
		child_group = frappe.get_doc("Tally Stock Group", "Sync Child Group")
		item = frappe.get_doc("Tally Item", "Sync Imported Item")
		ledger = frappe.get_doc("Tally Customer Ledger", "SYNC-LEDGER-001")

		self.assertEqual(run.status, "Completed")
		self.assertEqual(run.records_seen, 4)
		self.assertTrue(root_group.is_root)
		self.assertEqual(child_group.parent_stock_group, root_group.name)
		self.assertEqual(child_group.root_stock_group, root_group.name)
		self.assertEqual(child_group.depth, 1)
		self.assertEqual(child_group.full_path, "Sync Root Group > Sync Child Group")
		self.assertEqual(item.immediate_stock_group, child_group.name)
		self.assertEqual(item.root_stock_group, root_group.name)
		self.assertEqual(item.total_closing_balance, 12)
		self.assertEqual(ledger.ledger_name, "Sync Ledger Customer")

	def test_tally_master_sync_renames_existing_records_by_guid(self):
		group = self._create_product_group("Old GUID Group")
		item = self._create_item("Old GUID Item", group.name)
		frappe.db.set_value("Tally Stock Group", group.name, "tally_guid", "GROUP-RENAME-001")
		frappe.db.set_value("Tally Item", item.name, "tally_guid", "ITEM-RENAME-001")

		run = sync_tally_masters(
			{
				"stock_groups": [{"group_name": "New GUID Group", "tally_guid": "GROUP-RENAME-001", "is_active": 1}],
				"items": [
					{
						"item_name": "New GUID Item",
						"tally_guid": "ITEM-RENAME-001",
						"immediate_stock_group": "New GUID Group",
						"root_stock_group": "New GUID Group",
						"uom": "PCS",
						"is_active": 1,
					}
				],
			}
		)

		self.assertEqual(run.status, "Completed")
		self.assertFalse(frappe.db.exists("Tally Stock Group", "Old GUID Group"))
		self.assertFalse(frappe.db.exists("Tally Item", "Old GUID Item"))
		self.assertEqual(frappe.db.get_value("Tally Stock Group", "New GUID Group", "tally_guid"), "GROUP-RENAME-001")
		self.assertEqual(frappe.db.get_value("Tally Item", "New GUID Item", "immediate_stock_group"), "New GUID Group")

	def test_tally_master_sync_updates_stock_group_parent_move(self):
		self._create_product_group("Old Parent")
		self._create_product_group("New Parent")
		self._create_product_group("Moved Child")

		run = sync_tally_masters(
			{
				"stock_groups": [
					{"group_name": "Old Parent", "tally_guid": self._test_guid("GROUP", "Old Parent"), "is_active": 1},
					{"group_name": "New Parent", "tally_guid": self._test_guid("GROUP", "New Parent"), "is_active": 1},
					{
						"group_name": "Moved Child",
						"tally_guid": self._test_guid("GROUP", "Moved Child"),
						"source_parent_group": "New Parent",
						"is_active": 1,
					},
				],
			}
		)

		child = frappe.get_doc("Tally Stock Group", "Moved Child")
		self.assertEqual(run.status, "Completed")
		self.assertEqual(child.parent_stock_group, "New Parent")
		self.assertEqual(child.root_stock_group, "New Parent")
		self.assertEqual(child.depth, 1)

	def test_tally_ledger_guid_rename_updates_customer_client_code(self):
		frappe.get_doc(
			{
				"doctype": "Tally Customer Ledger",
				"client_code": "LEDGER-OLD-001",
				"ledger_name": "Ledger Rename Customer",
				"tally_guid": "LEDGER-RENAME-001",
				"is_active": 1,
			}
		).insert()
		customer = frappe.get_doc(
			{
				"doctype": "Customer",
				"customer_name": "Ledger Rename Customer",
				"business_legal_name": "Ledger Rename Customer",
				"mobile_number": "9000099001",
				"client_code": "LEDGER-OLD-001",
			}
		).insert()

		run = sync_tally_masters(
			{
				"customer_ledgers": [
					{
						"client_code": "LEDGER-NEW-001",
						"ledger_name": "Ledger Rename Customer",
						"tally_guid": "LEDGER-RENAME-001",
						"is_active": 1,
					}
				],
			}
		)

		self.assertEqual(run.status, "Completed")
		self.assertFalse(frappe.db.exists("Tally Customer Ledger", "LEDGER-OLD-001"))
		self.assertEqual(frappe.db.get_value("Customer", customer.name, "client_code"), "LEDGER-NEW-001")

	def test_stock_snapshot_uses_guid_pair_after_item_rename(self):
		group = self._create_product_group("Snapshot GUID Group")
		item = self._create_item("Snapshot Old Item", group.name)
		godown = self._create_godown("Snapshot GUID Godown")
		frappe.db.set_value("Tally Stock Group", group.name, "tally_guid", "GROUP-SNAPSHOT-001")
		frappe.db.set_value("Tally Item", item.name, "tally_guid", "ITEM-SNAPSHOT-001")
		frappe.db.set_value("Tally Godown", godown.name, "tally_guid", "GODOWN-SNAPSHOT-001")

		sync_stock_snapshots([{"item": item.name, "godown": godown.name, "quantity": 2, "uom": "PCS"}])
		sync_tally_masters(
			{
				"stock_groups": [{"group_name": group.name, "tally_guid": "GROUP-SNAPSHOT-001", "is_active": 1}],
				"items": [
					{
						"item_name": "Snapshot New Item",
						"tally_guid": "ITEM-SNAPSHOT-001",
						"immediate_stock_group": group.name,
						"root_stock_group": group.name,
						"uom": "PCS",
						"is_active": 1,
					}
				],
			}
		)
		sync_stock_snapshots([{"item": "Snapshot New Item", "godown": godown.name, "quantity": 7, "uom": "PCS"}])

		key = "ITEM-SNAPSHOT-001:GODOWN-SNAPSHOT-001"
		self.assertEqual(frappe.db.count("Tally Stock Snapshot", {"tally_snapshot_key": key}), 1)
		self.assertEqual(frappe.db.get_value("Tally Stock Snapshot", {"tally_snapshot_key": key}, "quantity"), 7)

	def test_voucher_without_reference_is_unmatched_and_not_reconciled(self):
		group = self._create_product_group("Unmatched Voucher Group")
		item = self._create_item("Unmatched Voucher Item", group.name)
		godown = self._create_godown("Unmatched Voucher Godown")
		frappe.db.set_value("Tally Item", item.name, "tally_guid", "ITEM-UNMATCHED-001")
		frappe.db.set_value("Tally Godown", godown.name, "tally_guid", "GODOWN-UNMATCHED-001")

		sync_tally_vouchers(
			[
				{
					"voucher_type": "Delivery Challan",
					"voucher_number": "DC-UNMATCHED-001",
					"party_client_code": "UNMATCHED-CUSTOMER",
					"lines": [{"item": item.name, "godown": godown.name, "quantity": 1}],
				}
			]
		)
		voucher = frappe.get_doc("Tally Voucher", "DC-UNMATCHED-001")
		run = run_reconciliation()

		self.assertEqual(voucher.reconciliation_state, "Unmatched")
		self.assertFalse(voucher.reconciled)
		self.assertEqual(run.records_seen, 0)

	def test_owner_manual_sync_actions_wrap_master_stock_and_reconciliation_jobs(self):
		product_group = self._create_product_group("Manual Sync PG")
		item = self._create_item("Manual Sync Item", product_group.name)
		branch_user = self._create_role_user("manual.sync.branch@example.com", "Branch Manager")

		masters_response = sync_masters_now(
			role="Owner",
			records={
				"units": [{"unit_name": "MANUAL-PCS", "symbol": "PCS", "is_active": 1}],
			},
		)
		stock_response = sync_stock_now(
			role="Owner",
			records=[{"item": item.name, "godown": "Manual Sync Godown", "quantity": 9}],
		)
		reconciliation_response = run_reconciliation_now(role="Owner")
		try:
			frappe.set_user(branch_user.name)
			branch_response = sync_stock_now(
				role="Branch Manager",
				records=[{"item": item.name, "godown": "Manual Sync Godown", "quantity": 9}],
			)
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(masters_response["success"])
		self.assertTrue(stock_response["success"])
		self.assertTrue(reconciliation_response["success"])
		self.assertFalse(branch_response["success"])
		self.assertEqual(masters_response["data"]["sync_type"], "Masters")
		self.assertEqual(stock_response["data"]["sync_type"], "Stock")
		self.assertEqual(reconciliation_response["data"]["sync_type"], "Reconciliation")

	def test_owner_admin_sync_actions_do_not_trust_claimed_role_parameter(self):
		branch_user = self._create_role_user("sync.branch@example.com", "Branch Manager")
		owner_user = self._create_role_user("sync.owner@example.com", "Owner")

		try:
			frappe.set_user(branch_user.name)
			branch_response = sync_masters_now(
				role="Owner",
				records={"units": [{"unit_name": "CLAIMED-OWNER-PCS", "symbol": "PCS", "is_active": 1}]},
			)
			frappe.set_user(owner_user.name)
			owner_response = sync_masters_now(
				role="Owner",
				records={"units": [{"unit_name": "REAL-OWNER-PCS", "symbol": "PCS", "is_active": 1}]},
			)
		finally:
			frappe.set_user("Administrator")

		self.assertFalse(branch_response["success"])
		self.assertIn("Owner/Admin", branch_response["error"]["message"])
		self.assertTrue(owner_response["success"])
		self.assertFalse(frappe.db.exists("Tally Unit", "CLAIMED-OWNER-PCS"))
		self.assertTrue(frappe.db.exists("Tally Unit", "REAL-OWNER-PCS"))

	def test_manual_admin_api_role_hints_are_optional_compatibility_metadata(self):
		for method in (sync_masters_now, sync_stock_now, sync_vouchers_now, import_stock_excel_now, run_reconciliation_now):
			parameters = inspect.signature(method).parameters
			self.assertIn("role", parameters, method.__name__)
			role_parameter = parameters["role"]
			self.assertIsNone(role_parameter.default, method.__name__)

	def test_owner_admin_manual_sync_actions_authorize_from_session_without_role_hint(self):
		product_group = self._create_product_group("Session Manual Sync PG")
		item = self._create_item("Session Manual Sync Item", product_group.name)
		self._create_godown("Session Manual Sync Godown")
		owner_user = self._create_role_user("session.sync.owner@example.com", "Owner")
		admin_user = self._create_role_user("session.sync.admin@example.com", "Admin")

		try:
			frappe.set_user(owner_user.name)
			masters_response = sync_masters_now(
				records={
					"units": [{"unit_name": "SESSION-OWNER-PCS", "symbol": "PCS", "is_active": 1}],
				}
			)
			stock_response = sync_stock_now(
				records=[{"item": item.name, "godown": "Session Manual Sync Godown", "quantity": 11}]
			)

			frappe.set_user(admin_user.name)
			voucher_response = sync_vouchers_now(
				records=[
					{
						"voucher_type": "Delivery Challan",
						"voucher_number": "DC-SESSION-ADMIN",
						"reference_number": "KE-26-05-9994",
						"party_client_code": "SESSION-CUSTOMER",
						"tracking_number": "TRACK-SESSION-ADMIN",
						"lines": [
							{
								"item": item.name,
								"godown": "Session Manual Sync Godown",
								"quantity": 1,
							}
						],
					}
				],
			)
			reconciliation_response = run_reconciliation_now()
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(masters_response["success"])
		self.assertTrue(stock_response["success"])
		self.assertTrue(voucher_response["success"])
		self.assertTrue(reconciliation_response["success"])
		self.assertTrue(frappe.db.exists("Tally Unit", "SESSION-OWNER-PCS"))
		self.assertTrue(frappe.db.exists("Tally Voucher", "DC-SESSION-ADMIN"))

	def test_branch_and_guest_users_cannot_forge_role_for_any_manual_admin_api(self):
		product_group = self._create_product_group("Forbidden Manual Sync PG")
		item = self._create_item("Forbidden Manual Sync Item", product_group.name)
		self._create_godown("Forbidden Manual Sync Godown")
		branch_manager = self._create_role_user("sync.branch.manager@example.com", "Branch Manager")
		branch_employee = self._create_role_user("sync.branch.employee@example.com", "Branch Employee")

		def call_all_admin_apis():
			api_calls = (
				(
					sync_masters_now,
					{
						"role": "Owner",
						"records": {"units": [{"unit_name": "FORGED-PCS", "symbol": "PCS", "is_active": 1}]},
					},
				),
				(
					sync_stock_now,
					{
						"role": "Owner",
						"records": [{"item": item.name, "godown": "Forbidden Manual Sync Godown", "quantity": 5}],
					},
				),
				(
					sync_vouchers_now,
					{
						"role": "Owner",
						"records": [
							{
								"voucher_type": "Delivery Challan",
								"voucher_number": "DC-FORGED-SYNC",
								"reference_number": "KE-26-05-9993",
								"party_client_code": "FORGED-CUSTOMER",
								"tracking_number": "TRACK-FORGED",
								"lines": [
									{
										"item": item.name,
										"godown": "Forbidden Manual Sync Godown",
										"quantity": 1,
									}
								],
							}
						],
					},
				),
				(run_reconciliation_now, {"role": "Owner"}),
				(import_stock_excel_now, {"file_url": "/files/missing-forbidden-stock.xlsx", "role": "Owner"}),
			)
			responses = []
			for method, kwargs in api_calls:
				try:
					responses.append(method(**kwargs))
				except TypeError as error:
					self.fail(f"{method.__name__} should accept compatibility role hints: {error}")
			return responses

		try:
			for user in (branch_manager.name, branch_employee.name, "Guest"):
				frappe.set_user(user)
				for response in call_all_admin_apis():
					self.assertFalse(response["success"], user)
					self.assertIn("Owner/Admin", response["error"]["message"], user)
		finally:
			frappe.set_user("Administrator")

		self.assertFalse(frappe.db.exists("Tally Unit", "FORGED-PCS"))
		self.assertFalse(frappe.db.exists("Tally Voucher", "DC-FORGED-SYNC"))

	def test_mixed_branch_role_does_not_remove_owner_admin_manual_api_access(self):
		product_group = self._create_product_group("Mixed Manual Sync PG")
		item = self._create_item("Mixed Manual Sync Item", product_group.name)
		self._create_godown("Mixed Manual Sync Godown")
		mixed_user = self._create_role_user("sync.mixed@example.com", "Owner")
		mixed_user.add_roles("Branch Manager")

		try:
			frappe.set_user(mixed_user.name)
			response = sync_stock_now(
				role="Branch Manager",
				records=[{"item": item.name, "godown": "Mixed Manual Sync Godown", "quantity": 7}],
			)
		finally:
			frappe.set_user("Administrator")

		self.assertTrue(response["success"])
		self.assertTrue(frappe.db.exists("Tally Stock Snapshot", f"{item.name}-Mixed Manual Sync Godown"))

	def test_scheduler_registers_five_minute_master_stock_and_reconciliation_jobs(self):
		five_minute_jobs = hooks.scheduler_events["cron"]["*/5 * * * *"]

		self.assertIn("kunal_enterprises.integrations.tally_postgres.import_masters", five_minute_jobs)
		self.assertIn("kunal_enterprises.integrations.tally_postgres.import_stock_snapshots", five_minute_jobs)
		self.assertIn("kunal_enterprises.integrations.tally_postgres.import_vouchers", five_minute_jobs)
		self.assertIn("kunal_enterprises.cron.reconciliation.run_reconciliation", five_minute_jobs)

	def test_voucher_sync_upserts_headers_and_lines_for_reconciliation(self):
		product_group = self._create_product_group("Voucher Sync PG")
		item = self._create_item("Voucher Sync Item", product_group.name)
		self._create_godown("Voucher Sync Godown")

		run = sync_tally_vouchers(
			[
				{
					"voucher_type": "Sales Invoice",
					"voucher_number": "SI-SYNC-001",
					"reference_number": "KE-26-05-9999",
					"party_client_code": "SYNC-CUSTOMER-001",
					"tracking_number": "TRACK-SYNC-001",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Voucher Sync Godown",
							"quantity": 2,
							"tracking_number": "TRACK-SYNC-001",
						}
					],
				},
				{
					"voucher_type": "Sales Invoice",
					"voucher_number": "SI-SYNC-001",
					"reference_number": "KE-26-05-9999",
					"party_client_code": "SYNC-CUSTOMER-001",
					"tracking_number": "TRACK-SYNC-001",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Voucher Sync Godown",
							"quantity": 4,
							"tracking_number": "TRACK-SYNC-001",
						}
					],
				},
			]
		)
		voucher = frappe.get_doc("Tally Voucher", "SI-SYNC-001")

		self.assertEqual(run.status, "Completed")
		self.assertEqual(run.records_seen, 2)
		self.assertEqual(run.records_processed, 2)
		self.assertEqual(voucher.reference_number, "KE-26-05-9999")
		self.assertEqual(len(voucher.lines), 1)
		self.assertEqual(voucher.lines[0].item, item.name)
		self.assertEqual(voucher.lines[0].quantity, 4)
		self.assertFalse(voucher.reconciled)

	def test_owner_manual_voucher_sync_action_wraps_voucher_job(self):
		product_group = self._create_product_group("Manual Voucher Sync PG")
		item = self._create_item("Manual Voucher Sync Item", product_group.name)
		self._create_godown("Manual Voucher Godown")

		response = sync_vouchers_now(
			role="Owner",
			records=[
				{
					"voucher_type": "Delivery Challan",
					"voucher_number": "DC-SYNC-001",
					"reference_number": "KE-26-05-9998",
					"party_client_code": "SYNC-CUSTOMER-002",
					"tracking_number": "TRACK-SYNC-002",
					"lines": [
						{
							"item": item.name,
							"godown": "Manual Voucher Godown",
							"quantity": 1,
						}
					],
				}
			],
		)

		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["sync_type"], "Vouchers")
		self.assertTrue(frappe.db.exists("Tally Voucher", "DC-SYNC-001"))

	def test_voucher_sync_logs_unknown_item_and_godown_lines_without_dropping_good_vouchers(self):
		product_group = self._create_product_group("Voucher Line Validation PG")
		item = self._create_item("Voucher Line Validation Item", product_group.name)
		self._create_godown("Voucher Line Validation Godown")

		run = sync_tally_vouchers(
			[
				{
					"voucher_type": "Sales Invoice",
					"voucher_number": "SI-SYNC-VALID-LINE",
					"reference_number": "KE-26-05-9997",
					"party_client_code": "SYNC-CUSTOMER-003",
					"tracking_number": "TRACK-SYNC-003",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Voucher Line Validation Godown",
							"quantity": 3,
						}
					],
				},
				{
					"voucher_type": "Sales Invoice",
					"voucher_number": "SI-SYNC-MISSING-ITEM",
					"reference_number": "KE-26-05-9996",
					"party_client_code": "SYNC-CUSTOMER-003",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": "Missing Voucher Item",
							"godown": "Voucher Line Validation Godown",
							"quantity": 1,
						}
					],
				},
				{
					"voucher_type": "Sales Invoice",
					"voucher_number": "SI-SYNC-MISSING-GODOWN",
					"reference_number": "KE-26-05-9995",
					"party_client_code": "SYNC-CUSTOMER-003",
					"voucher_date": "2026-05-19",
					"lines": [
						{
							"item": item.name,
							"godown": "Unknown Voucher Godown",
							"quantity": 1,
						}
					],
				},
			]
		)
		errors = frappe.get_all(
			"Tally Sync Error",
			filters={"sync_run": run.name},
			fields=["source_key", "error_message"],
			order_by="source_key asc",
		)

		self.assertEqual(run.status, "Completed With Errors")
		self.assertEqual(run.records_seen, 3)
		self.assertEqual(run.records_processed, 1)
		self.assertEqual(run.errors_count, 2)
		self.assertTrue(frappe.db.exists("Tally Voucher", "SI-SYNC-VALID-LINE"))
		self.assertFalse(frappe.db.exists("Tally Voucher", "SI-SYNC-MISSING-ITEM"))
		self.assertFalse(frappe.db.exists("Tally Voucher", "SI-SYNC-MISSING-GODOWN"))
		self.assertEqual(
			[row["source_key"] for row in errors],
			["SI-SYNC-MISSING-GODOWN", "SI-SYNC-MISSING-ITEM"],
		)
		self.assertIn("Tally Godown", errors[0]["error_message"])
		self.assertIn("Tally Item", errors[1]["error_message"])

	def test_stock_snapshot_sync_upserts_valid_rows_and_records_run(self):
		product_group = self._create_product_group("Sync PG")
		item = self._create_item("Sync Item", product_group.name)
		self._create_godown("Main Godown")

		run = sync_stock_snapshots(
			[
				{
					"item": item.name,
					"godown": "Main Godown",
					"quantity": 5,
					"uom": "PCS",
					"as_on_date": "2026-05-19",
					"source_company": "Kunal Test Company",
					"synced_at": "2026-05-19 12:00:00",
				},
				{
					"item": item.name,
					"godown": "Main Godown",
					"quantity": 8,
					"uom": "PCS",
					"as_on_date": "2026-05-19",
					"source_company": "Kunal Test Company",
					"synced_at": "2026-05-19 12:05:00",
				},
			]
		)

		snapshot = frappe.get_doc("Tally Stock Snapshot", f"{item.name}-Main Godown")
		self.assertEqual(run.status, "Completed")
		self.assertEqual(run.records_seen, 2)
		self.assertEqual(run.records_processed, 2)
		self.assertEqual(run.errors_count, 0)
		self.assertEqual(snapshot.quantity, 8)
		self.assertEqual(snapshot.source_sync_run, run.name)

	def test_stock_snapshot_sync_logs_bad_rows_without_dropping_good_rows(self):
		product_group = self._create_product_group("Sync Error PG")
		item = self._create_item("Sync Error Item", product_group.name)
		self._create_godown("Main Godown")

		run = sync_stock_snapshots(
			[
				{
					"item": item.name,
					"godown": "Main Godown",
					"quantity": 4,
					"uom": "PCS",
				},
				{
					"item": "Missing Item",
					"godown": "Main Godown",
					"quantity": 4,
					"uom": "PCS",
				},
			]
		)
		errors = frappe.get_all(
			"Tally Sync Error",
			filters={"sync_run": run.name},
			fields=["source_key", "error_message"],
		)

		self.assertEqual(run.status, "Completed With Errors")
		self.assertEqual(run.records_seen, 2)
		self.assertEqual(run.records_processed, 1)
		self.assertEqual(run.errors_count, 1)
		self.assertTrue(frappe.db.exists("Tally Stock Snapshot", f"{item.name}-Main Godown"))
		self.assertEqual(errors[0]["source_key"], "Missing Item:Main Godown")
		self.assertIn("does not exist", errors[0]["error_message"])

	def test_stock_snapshot_sync_logs_unknown_godown_rows(self):
		product_group = self._create_product_group("Sync Godown Error PG")
		item = self._create_item("Sync Godown Error Item", product_group.name)
		self._create_godown("Known Stock Godown")

		run = sync_stock_snapshots(
			[
				{
					"item": item.name,
					"godown": "Known Stock Godown",
					"quantity": 4,
					"uom": "PCS",
				},
				{
					"item": item.name,
					"godown": "Unknown Stock Godown",
					"quantity": 6,
					"uom": "PCS",
				},
			]
		)
		errors = frappe.get_all(
			"Tally Sync Error",
			filters={"sync_run": run.name},
			fields=["source_key", "error_message"],
		)

		self.assertEqual(run.status, "Completed With Errors")
		self.assertEqual(run.records_seen, 2)
		self.assertEqual(run.records_processed, 1)
		self.assertEqual(run.errors_count, 1)
		self.assertTrue(frappe.db.exists("Tally Stock Snapshot", f"{item.name}-Known Stock Godown"))
		self.assertFalse(frappe.db.exists("Tally Stock Snapshot", f"{item.name}-Unknown Stock Godown"))
		self.assertEqual(errors[0]["source_key"], f"{item.name}:Unknown Stock Godown")
		self.assertIn("Tally Godown", errors[0]["error_message"])

	def test_tally_stock_excel_parser_flattens_item_godown_rows(self):
		from kunal_enterprises.integrations.tally_stock_excel import parse_tally_stock_excel

		path = self._build_tally_stock_workbook(
			[
				("group", "Excel Import Group", 125),
				("item", "Excel Import Item A", 10),
				("godown", "Main Godown", 10),
				("item", "Excel Import Item B", 15),
				("godown", "Main Godown", 5),
				("godown", "Secondary Godown", 10),
				("group", "Ignored Group Total", 99),
				("item", "Excel Import Item C", 8),
				("godown", "Main Godown", 3),
				("batch", "Batch 1", "Main Godown", 5),
			]
		)

		rows = parse_tally_stock_excel(path)

		self.assertEqual(
			rows,
			[
				{
					"item": "Excel Import Item A",
					"godown": "Main Godown",
					"quantity": 10.0,
					"as_on_date": "2026-06-20",
				},
				{
					"item": "Excel Import Item B",
					"godown": "Main Godown",
					"quantity": 5.0,
					"as_on_date": "2026-06-20",
				},
				{
					"item": "Excel Import Item B",
					"godown": "Secondary Godown",
					"quantity": 10.0,
					"as_on_date": "2026-06-20",
				},
				{
					"item": "Excel Import Item C",
					"godown": "Main Godown",
					"quantity": 8.0,
					"as_on_date": "2026-06-20",
				},
			],
		)

	def test_tally_stock_excel_import_reuses_stock_snapshot_upsert(self):
		from kunal_enterprises.integrations.tally_stock_excel import import_tally_stock_excel_path

		product_group = self._create_product_group("Excel Import PG")
		item_a = self._create_item("Excel Import Sync Item A", product_group.name)
		item_b = self._create_item("Excel Import Sync Item B", product_group.name)
		self._create_godown("Excel Import Main Godown")
		self._create_godown("Excel Import Secondary Godown")

		first_path = self._build_tally_stock_workbook(
			[
				("group", "Excel Import Sync Group", 15),
				("item", item_a.name, 10),
				("godown", "Excel Import Main Godown", 10),
				("item", item_b.name, 5),
				("godown", "Excel Import Secondary Godown", 5),
			]
		)
		first_run = import_tally_stock_excel_path(first_path)

		second_path = self._build_tally_stock_workbook(
			[
				("group", "Excel Import Sync Group", 21),
				("item", item_a.name, 16),
				("godown", "Excel Import Main Godown", 16),
				("item", item_b.name, 5),
				("godown", "Excel Import Secondary Godown", 5),
			]
		)
		second_run = import_tally_stock_excel_path(second_path)

		snapshot_a = frappe.get_doc("Tally Stock Snapshot", f"{item_a.name}-Excel Import Main Godown")
		snapshot_b = frappe.get_doc("Tally Stock Snapshot", f"{item_b.name}-Excel Import Secondary Godown")
		self.assertEqual(first_run.status, "Completed")
		self.assertEqual(first_run.records_seen, 2)
		self.assertEqual(first_run.records_processed, 2)
		self.assertEqual(first_run.source_table, "tally_stock_excel")
		self.assertEqual(second_run.status, "Completed")
		self.assertEqual(second_run.records_seen, 2)
		self.assertEqual(snapshot_a.quantity, 16)
		self.assertEqual(snapshot_a.as_on_date.strftime("%Y-%m-%d"), "2026-06-20")
		self.assertEqual(snapshot_a.source_sync_run, second_run.name)
		self.assertEqual(snapshot_b.quantity, 5)

	def test_owner_manual_excel_stock_import_action_uses_controlled_file_record(self):
		product_group = self._create_product_group("Manual Excel Import PG")
		item = self._create_item("Manual Excel Import Item", product_group.name)
		self._create_godown("Manual Excel Import Godown")
		file_doc = self._save_tally_stock_workbook_file(
			[
				("group", "Manual Excel Import Group", 13),
				("item", item.name, 13),
				("godown", "Manual Excel Import Godown", 13),
			],
			"manual-excel-stock-import.xlsx",
		)
		owner_user = self._create_role_user("excel.import.owner@example.com", "Owner")

		try:
			frappe.set_user(owner_user.name)
			response = import_stock_excel_now(file_doc.file_url)
		finally:
			frappe.set_user("Administrator")

		snapshot = frappe.get_doc("Tally Stock Snapshot", f"{item.name}-Manual Excel Import Godown")
		self.assertTrue(response["success"])
		self.assertEqual(response["data"]["sync_type"], "Stock")
		self.assertEqual(snapshot.quantity, 13)

	def test_tally_stock_excel_import_logs_errors_with_excel_source(self):
		from kunal_enterprises.integrations.tally_stock_excel import EXCEL_STOCK_SOURCE_TABLE, import_tally_stock_excel_path

		product_group = self._create_product_group("Excel Import Error PG")
		item = self._create_item("Excel Import Error Item", product_group.name)
		self._create_godown("Excel Import Error Godown")
		path = self._build_tally_stock_workbook(
			[
				("group", "Excel Import Error Group", 9),
				("item", item.name, 4),
				("godown", "Excel Import Error Godown", 4),
				("item", "Missing Excel Import Item", 5),
				("godown", "Excel Import Error Godown", 5),
			]
		)

		run = import_tally_stock_excel_path(path)
		error = frappe.get_all(
			"Tally Sync Error",
			filters={"sync_run": run.name},
			fields=["source_table", "source_key", "error_message"],
		)[0]

		self.assertEqual(run.status, "Completed With Errors")
		self.assertEqual(run.records_seen, 2)
		self.assertEqual(run.records_processed, 1)
		self.assertEqual(run.errors_count, 1)
		self.assertEqual(error["source_table"], EXCEL_STOCK_SOURCE_TABLE)
		self.assertEqual(error["source_key"], "Missing Excel Import Item:Excel Import Error Godown")
		self.assertIn("Tally Item", error["error_message"])

	def test_dev_stock_snapshot_seed_requires_explicit_confirmation(self):
		with self.assertRaises(frappe.ValidationError):
			seed_dev_stock_snapshots()

	def test_dev_stock_snapshot_seed_rows_are_deterministic_and_cover_godowns(self):
		items = [
			{"name": "Seed Item A", "uom": "PCS"},
			{"name": "Seed Item B", "uom": "MTR"},
			{"name": "Seed Item C", "uom": "NOS"},
		]
		godowns = ["Goshamahal", "Kukatpally", "Main Location"]

		first_rows = _build_dev_stock_snapshot_rows(
			items=items,
			godowns=godowns,
			max_godowns_per_item=2,
			zero_stock_ratio=1,
			as_on_date="2026-05-21",
		)
		second_rows = _build_dev_stock_snapshot_rows(
			items=items,
			godowns=godowns,
			max_godowns_per_item=2,
			zero_stock_ratio=1,
			as_on_date="2026-05-21",
		)

		self.assertEqual(first_rows, second_rows)
		self.assertGreaterEqual(len(first_rows), len(items))
		self.assertLessEqual(len(first_rows), len(items) * 2)
		self.assertEqual({row["quantity"] for row in first_rows}, {0})
		self.assertTrue({row["godown"] for row in first_rows}.issubset(set(godowns)))
		self.assertEqual({row["as_on_date"] for row in first_rows}, {"2026-05-21"})

	def _create_product_group(self, group_name):
		return frappe.get_doc(
			{
				"doctype": "Tally Stock Group",
				"group_name": group_name,
				"tally_guid": self._test_guid("GROUP", group_name),
				"is_root": 1,
				"depth": 0,
				"full_path": group_name,
				"is_active": 1,
			}
		).insert()

	def _create_item(self, item_name, root_stock_group):
		return frappe.get_doc(
			{
				"doctype": "Tally Item",
				"item_name": item_name,
				"tally_guid": self._test_guid("ITEM", item_name),
				"root_stock_group": root_stock_group,
				"uom": "PCS",
				"total_closing_balance": 0,
				"is_active": 1,
			}
		).insert()

	def _create_godown(self, godown_name, is_active=1):
		if frappe.db.exists("Tally Godown", godown_name):
			frappe.db.set_value(
				"Tally Godown", godown_name, {"is_active": is_active, "tally_guid": self._test_guid("GODOWN", godown_name)}
			)
			return frappe.get_doc("Tally Godown", godown_name)
		return frappe.get_doc(
			{
				"doctype": "Tally Godown",
				"godown_name": godown_name,
				"tally_guid": self._test_guid("GODOWN", godown_name),
				"is_active": is_active,
			}
		).insert()

	def _test_guid(self, prefix, value):
		return f"TEST-{prefix}-{value}".upper().replace(" ", "-")

	def _build_tally_stock_workbook(self, rows):
		import os
		import tempfile

		from openpyxl import Workbook
		from openpyxl.styles import Font

		workbook = Workbook()
		sheet = workbook.active
		sheet.title = "Stock Summary"
		sheet.cell(row=1, column=1, value="Particulars - 20.06.26")
		sheet.cell(row=1, column=3, value="Quantity")

		for row_number, row in enumerate(rows, start=2):
			row_type = row[0]
			if row_type == "group":
				sheet.cell(row=row_number, column=1, value=row[1])
				sheet.cell(row=row_number, column=3, value=row[2])
				sheet.cell(row=row_number, column=1).font = Font(bold=True)
			elif row_type == "item":
				sheet.cell(row=row_number, column=1, value=row[1])
				sheet.cell(row=row_number, column=3, value=row[2])
				sheet.cell(row=row_number, column=1).font = Font(italic=True)
			elif row_type == "godown":
				sheet.cell(row=row_number, column=2, value=row[1])
				sheet.cell(row=row_number, column=3, value=row[2])
			elif row_type == "batch":
				sheet.cell(row=row_number, column=1, value=row[1])
				sheet.cell(row=row_number, column=2, value=row[2])
				sheet.cell(row=row_number, column=3, value=row[3])

		handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		handle.close()
		workbook.save(handle.name)
		self.addCleanup(lambda: os.path.exists(handle.name) and os.remove(handle.name))
		return handle.name

	def _save_tally_stock_workbook_file(self, rows, filename):
		import os

		from frappe.utils.file_manager import get_file_path, save_file

		path = self._build_tally_stock_workbook(rows)
		with open(path, "rb") as workbook_file:
			file_doc = save_file(filename, workbook_file.read(), None, None, is_private=1)
		file_path = get_file_path(file_doc.file_url)
		self.addCleanup(lambda: os.path.exists(file_path) and os.remove(file_path))
		return file_doc

	def _create_role_user(self, email, role):
		user = frappe.get_doc(
			{
				"doctype": "User",
				"email": email,
				"first_name": email.split("@")[0],
				"enabled": 1,
				"send_welcome_email": 0,
			}
		).insert(ignore_permissions=True)
		user.add_roles(role)
		return user
